import os
import datetime
import pandas as pd
import gspread
from oauth2client.service_account import ServiceAccountCredentials
from openpyxl import load_workbook
import matplotlib.pyplot as plt
from fpdf import FPDF
import re

# --- COLORES PARA LA TERMINAL (ESTÉTICA) ---
class CTerm:
    CYAN = '\033[96m'
    BLUE = '\033[94m'
    GREEN = '\033[92m'
    YELLOW = '\033[93m'
    RED = '\033[91m'
    BOLD = '\033[1m'
    END = '\033[0m'

# --- CONFIGURACION DINÁMICA DE RUTAS ---
# Sube un nivel desde /bots hasta la raíz de SunHaven
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

ARCHIVO_MASTER = os.path.join(BASE_DIR, 'data', 'inputs', 'Master_Bonos.xlsx')
ARCHIVO_CREDENCIALES = os.path.join(BASE_DIR, 'config', 'sunhaven-credentials.json')
CARPETA_OUTPUTS = os.path.join(BASE_DIR, 'data', 'outputs')

# PON AQUÍ EL NOMBRE EXACTO DE TU GOOGLE SHEET:
NOMBRE_DEL_GOOGLE_SHEET = "SUNHAVEN_KAIZEN (Respuestas)" 

EXCEPCIONES_KAIZEN = ["Yareli Yamile Luquin Puga", "Blanca Aracely Figueroa Marroquín"]

# --- PALETA EJECUTIVA SUN HAVEN ---
C_NAVY = (31, 58, 82)
C_SUN = (211, 84, 0)
C_LIGHT_BG = (245, 247, 248)
C_DARK = (44, 62, 80)
C_SEC = (189, 195, 199)

def limpiar_texto(texto):
    if not isinstance(texto, str): return ""
    texto = re.sub(r'\s+', ' ', texto).strip()
    if len(texto) > 1:
        texto = texto[0].upper() + texto[1:]
    return texto

def normalizar_nombre(nombre):
    if pd.isna(nombre): return ""
    return " ".join(str(nombre).strip().upper().split())

class KaizenPDF(FPDF):
    def header(self):
        if self.page_no() == 1: return
        self.set_fill_color(*C_NAVY)
        self.rect(0, 0, 210, 26, 'F')
        self.set_fill_color(*C_SUN)
        self.rect(0, 26, 210, 1.5, 'F')
        self.set_y(9)
        self.set_font('Helvetica', 'B', 16)
        self.set_text_color(255, 255, 255)
        self.cell(0, 8, 'REPORTE EJECUTIVO DE MEJORA CONTINUA (KAIZEN)', 0, 1, 'C')
        self.ln(12)

    def footer(self):
        if self.page_no() == 1: return
        self.set_y(-15)
        self.set_font('Helvetica', 'I', 8)
        self.set_text_color(*C_NAVY)
        self.set_draw_color(*C_SUN)
        self.line(10, 282, 200, 282)
        
        self.set_x(10)
        self.cell(190, 10, f'Página {self.page_no()}', 0, 0, 'C')
        
        self.set_x(10)
        self.cell(190, 10, 'Ing. Larry Beresford', 0, 0, 'R')

    def cover_page(self, mes_nombre, anio):
        self.add_page()
        self.set_fill_color(*C_NAVY)
        self.rect(0, 0, 210, 100, 'F')
        self.set_fill_color(*C_SUN)
        self.rect(0, 100, 210, 3, 'F')
        self.set_y(40)
        self.set_font('Helvetica', 'B', 28)
        self.set_text_color(255, 255, 255)
        self.cell(0, 10, 'SUN HAVEN', 0, 1, 'C')
        self.set_font('Helvetica', '', 12)
        self.set_text_color(220, 220, 220)
        self.cell(0, 8, 'INNOVACIÓN Y MEJORA CONTINUA', 0, 1, 'C')
        self.set_y(140)
        self.set_font('Helvetica', 'B', 22)
        self.set_text_color(*C_DARK)
        self.cell(0, 10, 'REPORTE EJECUTIVO', 0, 1, 'C')
        self.set_font('Helvetica', 'B', 26)
        self.set_text_color(*C_SUN)
        self.cell(0, 12, 'KAIZEN', 0, 1, 'C')
        self.set_y(220)
        self.set_font('Helvetica', 'B', 12)
        self.set_text_color(*C_NAVY)
        self.cell(0, 8, f'PERIODO EVALUADO: {mes_nombre.upper()} {anio}', 0, 1, 'C')
        
        self.set_y(260)
        self.set_font('Helvetica', '', 11)
        self.set_text_color(*C_DARK)
        self.cell(0, 6, 'Elaborado por:', 0, 1, 'C')
        self.set_font('Helvetica', 'B', 11)
        self.cell(0, 6, 'Ing. Larry Beresford', 0, 1, 'C')

def main():
    print(f"\n{CTerm.CYAN}{CTerm.BOLD}============================================================{CTerm.END}")
    print(f"{CTerm.CYAN}{CTerm.BOLD}   SISTEMA KAIZEN - EXTRACCIÓN Y AUDITORÍA{CTerm.END}")
    print(f"{CTerm.CYAN}{CTerm.BOLD}============================================================{CTerm.END}\n")
    
    # 1. PEDIR FECHA
    try:
        mes_actual = datetime.date.today().month
        anio_actual = datetime.date.today().year
        entrada_mes = input(f"{CTerm.YELLOW}Ingrese el MES a evaluar (1-12) [Enter para {mes_actual}]: {CTerm.END}").strip()
        mes = int(entrada_mes) if entrada_mes else mes_actual
        entrada_anio = input(f"{CTerm.YELLOW}Ingrese el AÑO (ej. 2026) [Enter para {anio_actual}]: {CTerm.END}").strip()
        anio = int(entrada_anio) if entrada_anio else anio_actual
    except ValueError:
        print(f"\n{CTerm.RED}[ERROR] Ingresa números válidos.{CTerm.END}")
        return

    meses_nombres = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
    nombre_mes = meses_nombres[mes - 1]
    
    mes_prev = mes - 1 if mes > 1 else 12
    anio_prev = anio if mes > 1 else anio - 1
    nombre_mes_prev = meses_nombres[mes_prev - 1]

    fecha_castigo = datetime.date(anio, mes, 25) 

    # 2. CONECTAR A GOOGLE SHEETS
    print(f"\n{CTerm.BLUE}[1/4] Conectando a la nube de Google para buscar propuestas de {nombre_mes} {anio}...{CTerm.END}")
    try:
        alcance = ['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive']
        credenciales = ServiceAccountCredentials.from_json_keyfile_name(ARCHIVO_CREDENCIALES, alcance)
        cliente = gspread.authorize(credenciales)
        hoja = cliente.open(NOMBRE_DEL_GOOGLE_SHEET).sheet1
        datos_kaizen = hoja.get_all_records()
    except Exception as e:
        print(f"{CTerm.RED}[ERROR] No se pudo conectar a Google Sheets: {e}{CTerm.END}")
        return

    # 3. PROCESAR RESPUESTAS
    print(f"{CTerm.BLUE}[2/4] Procesando datos y cruzando con la plantilla de Master Bonos...{CTerm.END}")
    nombres_participaron_actual = set()
    nombres_participaron_prev = set()
    propuestas_limpias = []

    if datos_kaizen:
        df_kaizen = pd.DataFrame(datos_kaizen)
        
        if 'Marca temporal' in df_kaizen.columns:
            df_kaizen['Marca temporal'] = pd.to_datetime(df_kaizen['Marca temporal'], dayfirst=True, errors='coerce')
        
        for index, row in df_kaizen.iterrows():
            fecha_obj = row.get('Marca temporal')
            if pd.isna(fecha_obj):
                continue
            
            mes_form = fecha_obj.month
            anio_form = fecha_obj.year
            
            nombre_raw = str(row.get('Colaborador', '')).strip()
            propuesta_raw = str(row.get('Propuesta de mejora', '')).strip()
            
            if not nombre_raw or nombre_raw.lower() == 'nan':
                continue

            if mes_form == mes and anio_form == anio:
                nombres_participaron_actual.add(normalizar_nombre(nombre_raw))
                propuestas_limpias.append({
                    'nombre': nombre_raw,
                    'propuesta': limpiar_texto(propuesta_raw),
                    'fecha': fecha_obj.strftime('%d/%m/%Y')
                })
            elif mes_form == mes_prev and anio_form == anio_prev:
                nombres_participaron_prev.add(normalizar_nombre(nombre_raw))

    # 4. LEER MASTER BONOS
    try:
        df_personal = pd.read_excel(ARCHIVO_MASTER, sheet_name='Listado de personal')
        df_enfermeria = df_personal[df_personal['Departamento'].astype(str).str.contains('Enfermer', case=False, na=False)]
        nombres_obligados = [str(row['Nombre completo']).strip() for _, row in df_enfermeria.iterrows() if str(row['Nombre completo']).strip() not in EXCEPCIONES_KAIZEN and str(row['Nombre completo']).strip().lower() != 'nan']
    except Exception as e:
        print(f"{CTerm.RED}[ERROR] Leyendo Master Bonos: {e}{CTerm.END}")
        return

    lista_faltantes = [emp for emp in nombres_obligados if normalizar_nombre(emp) not in nombres_participaron_actual]

    # 5. INYECTAR PENALIZACIONES
    print(f"{CTerm.BLUE}[3/4] Inyectando {len(lista_faltantes)} penalizaciones en la bitácora...{CTerm.END}")
    try:
        wb = load_workbook(ARCHIVO_MASTER)
        ws = wb['Auditorías'] if 'Auditorías' in wb.sheetnames else wb['BITACORA']
        fila_actual = 2
        max_id = 0
        while ws.cell(row=fila_actual, column=1).value is not None:
            try:
                val = int(ws.cell(row=fila_actual, column=1).value)
                if val > max_id: max_id = val
            except: pass
            fila_actual += 1

        for faltante in lista_faltantes:
            max_id += 1
            ws.cell(row=fila_actual, column=1, value=max_id)
            ws.cell(row=fila_actual, column=2, value=fecha_castigo).number_format = 'YYYY-MM-DD'
            ws.cell(row=fila_actual, column=3, value=faltante)
            ws.cell(row=fila_actual, column=4, value="Falla Admin/Celular")
            ws.cell(row=fila_actual, column=5, value="No presentó Kaizen del mes")
            fila_actual += 1
            
        wb.save(ARCHIVO_MASTER)
        print(f"      {CTerm.GREEN}[ÉXITO] Penalizaciones guardadas con éxito en Master_Bonos.xlsx{CTerm.END}")
    except Exception as e:
        print(f"{CTerm.RED}[ERROR] No se pudo escribir en el Excel: {e}{CTerm.END}")

    # 6. CREAR GRÁFICA
    total_obligados = len(nombres_obligados)
    participantes_count = total_obligados - len(lista_faltantes)
    faltantes_count = len(lista_faltantes)
    participantes_count_prev = len([emp for emp in nombres_obligados if normalizar_nombre(emp) in nombres_participaron_prev])
    faltantes_count_prev = total_obligados - participantes_count_prev
    
    # Asegurar que la gráfica se guarda en la carpeta tools temporal
    ruta_grafica = os.path.join(BASE_DIR, 'tools', 'grafico_kaizen.png')
    
    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(10, 4.5))
    ax1.pie([participantes_count_prev, faltantes_count_prev], labels=['Participaron', 'No Participaron'], autopct='%1.1f%%', colors=['#1F3A52', '#D35400'], startangle=90, textprops={'fontsize': 10, 'color': 'white', 'weight': 'bold'})
    ax1.set_title(f"{nombre_mes_prev.upper()} {anio_prev}", fontweight='bold', color='#2C3E50', fontsize=12)
    ax2.pie([participantes_count, faltantes_count], labels=['Participaron', 'No Participaron'], autopct='%1.1f%%', colors=['#1F3A52', '#D35400'], startangle=90, textprops={'fontsize': 10, 'color': 'white', 'weight': 'bold'})
    ax2.set_title(f"{nombre_mes.upper()} {anio}", fontweight='bold', color='#2C3E50', fontsize=12)
    fig.legend(['Participaron', 'No Participaron'], loc="lower center", bbox_to_anchor=(0.5, 0.0), ncol=2, fontsize=10)
    plt.tight_layout()
    plt.savefig(ruta_grafica, transparent=True, dpi=300, bbox_inches='tight')
    plt.close()

    # 7. GENERAR PDF EJECUTIVO EN SU CARPETA
    print(f"{CTerm.BLUE}[4/4] Generando Reporte PDF Ejecutivo de Kaizen...{CTerm.END}")
    
    nombre_carpeta_mes = f"{mes:02d}_{nombre_mes}_{anio}"  
    ruta_carpeta_mes = os.path.join(CARPETA_OUTPUTS, nombre_carpeta_mes)
    os.makedirs(ruta_carpeta_mes, exist_ok=True)
    
    pdf = KaizenPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.cover_page(nombre_mes, anio)
    
    # --- PÁGINA DE ESTADÍSTICAS ---
    pdf.add_page()
    pdf.set_font('Helvetica', 'B', 12)
    pdf.set_text_color(*C_NAVY)
    pdf.cell(0, 8, "1. ESTADÍSTICAS DE PARTICIPACIÓN Y TENDENCIA", 0, 1, 'L')
    pdf.set_draw_color(*C_SUN)
    pdf.line(10, pdf.get_y(), 200, pdf.get_y())
    pdf.ln(5)
    
    pdf.set_font('Helvetica', '', 11)
    pdf.set_text_color(*C_DARK)
    pdf.multi_cell(0, 6, f"Comparativa de participación para los {total_obligados} colaboradores referentes al departamento de enfermería:\n\n- En el mes anterior ({nombre_mes_prev}) participaron {participantes_count_prev} y no participaron {faltantes_count_prev}.\n- En el mes evaluado ({nombre_mes}) participaron {participantes_count} y no participaron {faltantes_count}.")
    
    if os.path.exists(ruta_grafica):
        pdf.ln(5)
        y_graph = pdf.get_y()
        pdf.image(ruta_grafica, x=15, y=y_graph, w=180) 
        os.remove(ruta_grafica)
        pdf.set_y(y_graph + 85) 

    # --- LISTADO EN DOS COLUMNAS ---
    pdf.ln(5)
    
    participaron_list = sorted([emp for emp in nombres_obligados if emp not in lista_faltantes])
    faltantes_list = sorted(lista_faltantes)
    
    pdf.set_font('Helvetica', 'B', 10)
    pdf.set_text_color(*C_NAVY)
    pdf.cell(95, 6, "Colaboradores que SÍ participaron:", 0, 0, 'L')
    pdf.set_text_color(*C_SUN)
    pdf.cell(95, 6, "Colaboradores que NO participaron:", 0, 1, 'L')
    
    pdf.set_font('Helvetica', '', 9)
    pdf.set_text_color(*C_DARK)
    
    max_len = max(len(participaron_list), len(faltantes_list))
    for i in range(max_len):
        name_part = participaron_list[i] if i < len(participaron_list) else ""
        name_falt = faltantes_list[i] if i < len(faltantes_list) else ""
        
        text_part = f"   - {name_part}" if name_part else ""
        text_falt = f"   - {name_falt}" if name_falt else ""
        
        pdf.cell(95, 5, text_part, 0, 0, 'L')
        pdf.cell(95, 5, text_falt, 0, 1, 'L')
        
    pdf.ln(10)

    # --- PÁGINA DE PROPUESTAS ---
    pdf.add_page()
    pdf.set_font('Helvetica', 'B', 12)
    pdf.set_text_color(*C_NAVY)
    pdf.cell(0, 8, "2. PROPUESTAS DE MEJORA", 0, 1, 'L')
    pdf.line(10, pdf.get_y(), 200, pdf.get_y())
    pdf.ln(5)
    
    if propuestas_limpias:
        for prop in sorted(propuestas_limpias, key=lambda x: x['nombre']):
            y_inicio = pdf.get_y()
            pdf.set_fill_color(*C_LIGHT_BG)
            pdf.set_draw_color(*C_SEC)
            
            if y_inicio > 240:
                pdf.add_page()
                y_inicio = pdf.get_y()

            pdf.set_font('Helvetica', 'B', 11)
            pdf.set_text_color(*C_NAVY)
            pdf.cell(130, 8, f" Colaborador(a): {prop['nombre']}", 'L T', 0, 'L', fill=True)
            pdf.set_font('Helvetica', 'I', 9)
            pdf.cell(60, 8, f"Fecha de envío: {prop['fecha']} ", 'T R', 1, 'R', fill=True)
            pdf.set_font('Helvetica', '', 10)
            pdf.set_text_color(*C_DARK)
            pdf.multi_cell(0, 6, f" Propuesta:\n {prop['propuesta']}\n", 'L B R', 'L', fill=True)
            pdf.ln(4)
    else:
        pdf.set_font('Helvetica', '', 11)
        pdf.cell(0, 6, "No hay propuestas registradas para este periodo.", 0, 1)

    nombre_pdf = f"Reporte_Kaizen_{mes:02d}_{anio}.pdf"
    ruta_guardado = os.path.join(ruta_carpeta_mes, nombre_pdf)
    pdf.output(ruta_guardado)
    
    print(f"\n{CTerm.GREEN}{CTerm.BOLD}[ÉXITO TOTAL] Reporte guardado en: data/outputs/{nombre_carpeta_mes}/{nombre_pdf}{CTerm.END}\n")

if __name__ == "__main__":
    main()