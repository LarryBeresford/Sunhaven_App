import pandas as pd
import datetime
import os
from openpyxl import load_workbook
import matplotlib.pyplot as plt
from matplotlib.ticker import MaxNLocator
from fpdf import FPDF

# --- COLORES PARA LA TERMINAL (ESTÉTICA) ---
class CTerm:
    CYAN = '\033[96m'
    BLUE = '\033[94m'
    GREEN = '\033[92m'
    YELLOW = '\033[93m'
    RED = '\033[91m'
    BOLD = '\033[1m'
    END = '\033[0m'

# --- CONFIGURACION ---
CARPETA_ACTUAL = os.path.dirname(os.path.abspath(__file__))
ARCHIVO_MASTER = os.path.join(CARPETA_ACTUAL, 'Master_Bonos.xlsx')
ARCHIVO_ASISTENCIA = os.path.join(CARPETA_ACTUAL, 'Asistencia_abril.xlsx')

EXCEPCIONES_DASHBOARD = ["Yareli Yamile Luquin Puga", "Blanca Aracely Figueroa Marroquín", "Guadalupe Georgia Lopez Ceja"]
TURNO_NOCHE = ["Jaqueline Hernández Sosa", "Silvia Rodríguez Reynaga", "Consuelo Ceja Liborio", "Guadalupe Georgia Lopez Ceja"]
CHECADORES_ESPECIALES = ["CESAR", "MONI", "MarthaCastro", "HUGO"]

# --- PALETA EJECUTIVA SUN HAVEN (Basada en la Web) ---
C_NAVY = (31, 58, 82)         
C_SUN = (211, 84, 0)          
C_LIGHT_BG = (245, 247, 248)  
C_DARK = (44, 62, 80)         
C_SEC = (189, 195, 199)       

def normalizar_nombre(nombre):
    if pd.isna(nombre): return ""
    return " ".join(str(nombre).split())

def abreviar_nombre(nombre_completo):
    partes = str(nombre_completo).split()
    if len(partes) > 1:
        return f"{partes[0]} {partes[1][0]}."
    return partes[0] if partes else ""

def encontrar_fila_inicio_y_ultimo_id(hoja):
    fila = 2
    max_id = 0
    while True:
        val_id = hoja.cell(row=fila, column=1).value
        if val_id is None or str(val_id).strip() == "": return fila, max_id
        try:
            if int(val_id) > max_id: max_id = int(val_id)
        except: pass
        fila += 1

def time_to_mins(hora_str):
    h, m = map(int, hora_str.split(':'))
    return h * 60 + m

class PremiumPDF(FPDF):
    def header(self):
        if self.page_no() == 1: return
        self.set_fill_color(*C_NAVY)
        self.rect(0, 0, 210, 26, 'F')
        self.set_fill_color(*C_SUN)
        self.rect(0, 26, 210, 1.5, 'F')
        self.set_y(9)
        self.set_font('Helvetica', 'B', 16)
        self.set_text_color(255, 255, 255)
        self.cell(0, 8, 'REPORTE EJECUTIVO DE BONOS E INCIDENCIAS', 0, 1, 'C')
        self.ln(12)

    def footer(self):
        if self.page_no() == 1: return
        self.set_y(-15)
        self.set_font('Helvetica', 'I', 8)
        self.set_text_color(*C_NAVY)
        self.set_draw_color(*C_SUN)
        self.line(10, 282, 200, 282)
        self.set_x(10)
        self.cell(190, 10, f'Página {self.page_no()}', 0, 0, 'C')
        self.set_x(10)
        self.cell(190, 10, 'Ing. Larry Beresford', 0, 0, 'R')

    def chapter_title(self, title):
        self.set_font('Helvetica', 'B', 12)
        self.set_text_color(*C_NAVY)
        self.cell(0, 8, title.upper(), 0, 1, 'L')
        self.set_draw_color(*C_SUN)
        self.set_line_width(0.5)
        self.line(self.get_x(), self.get_y(), 200, self.get_y())
        self.ln(4)
        
    def cover_page(self, mes_nombre, anio):
        self.add_page()
        self.set_fill_color(*C_NAVY)
        self.rect(0, 0, 210, 100, 'F')
        self.set_fill_color(*C_SUN)
        self.rect(0, 100, 210, 3, 'F')
        self.set_y(40)
        self.set_font('Helvetica', 'B', 28)
        self.set_text_color(255, 255, 255)
        self.cell(0, 10, 'SUN HAVEN', 0, 1, 'C')
        self.set_font('Helvetica', '', 12)
        self.set_text_color(220, 220, 220)
        self.cell(0, 8, 'CASA DE DESCANSO PARA ADULTOS MAYORES', 0, 1, 'C')
        self.set_y(140)
        self.set_font('Helvetica', 'B', 22)
        self.set_text_color(*C_DARK)
        self.cell(0, 10, 'REPORTE EJECUTIVO', 0, 1, 'C')
        self.set_font('Helvetica', 'B', 26)
        self.set_text_color(*C_SUN)
        self.cell(0, 12, 'BONOS', 0, 1, 'C')
        self.set_y(220)
        self.set_font('Helvetica', 'B', 12)
        self.set_text_color(*C_NAVY)
        self.cell(0, 8, f'PERIODO EVALUADO: {mes_nombre.upper()} {anio}', 0, 1, 'C')
        self.set_y(260)
        self.set_font('Helvetica', '', 11)
        self.set_text_color(*C_DARK)
        self.cell(0, 6, 'Elaborado por:', 0, 1, 'C')
        self.set_font('Helvetica', 'B', 11)
        self.cell(0, 6, 'Ing. Larry Beresford', 0, 1, 'C')

def main():
    print(f"\n{CTerm.CYAN}{CTerm.BOLD}============================================================{CTerm.END}")
    print(f"{CTerm.CYAN}{CTerm.BOLD}   SISTEMA DE AUDITORÍA Y BONOS - SUN HAVEN{CTerm.END}")
    print(f"{CTerm.CYAN}{CTerm.BOLD}============================================================{CTerm.END}\n")
    
    # --- INTERFAZ DE USUARIO: SELECCIÓN DE MES Y AÑO ---
    try:
        mes_actual = datetime.date.today().month
        anio_actual = datetime.date.today().year
        
        entrada_mes = input(f"{CTerm.YELLOW}Ingrese el número del MES a evaluar (1-12) [Enter para mes actual ({mes_actual})]: {CTerm.END}").strip()
        mes = int(entrada_mes) if entrada_mes else mes_actual
        
        entrada_anio = input(f"{CTerm.YELLOW}Ingrese el AÑO a evaluar (ej. 2026) [Enter para año actual ({anio_actual})]: {CTerm.END}").strip()
        anio = int(entrada_anio) if entrada_anio else anio_actual
        
        if not (1 <= mes <= 12):
            print(f"\n{CTerm.RED}[ERROR] El mes debe ser un número entre 1 y 12. Inténtalo de nuevo.{CTerm.END}")
            return
            
    except ValueError:
        print(f"\n{CTerm.RED}[ERROR] Entrada inválida. Por favor ingresa solo números.{CTerm.END}")
        return

    meses_nombres = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
    nombre_mes_actual = meses_nombres[mes - 1]
    
    print(f"\n{CTerm.BLUE}[INFO] Iniciando procesamiento estricto para: {nombre_mes_actual.upper()} {anio}...{CTerm.END}")

    # 1. CARGAR NOMBRES
    try:
        df_personal = pd.read_excel(ARCHIVO_MASTER, sheet_name='Listado de personal')
        mapa_nombres_enfermeria = {normalizar_nombre(row['Nombre checador']): normalizar_nombre(row['Nombre completo']) for _, row in df_personal.iterrows() if normalizar_nombre(row['Nombre checador']) and 'Enfermer' in normalizar_nombre(row['Departamento'])}
        mapa_nombres_especiales = {normalizar_nombre(row['Nombre checador']): normalizar_nombre(row['Nombre completo']) for _, row in df_personal.iterrows() if normalizar_nombre(row['Nombre checador']) in CHECADORES_ESPECIALES}
    except Exception as e:
        print(f"{CTerm.RED}[ERROR] No se pudo leer el 'Listado de personal': {e}{CTerm.END}")
        return

    # 2. LEER ASISTENCIA
    try:
        print(f"{CTerm.BLUE}[INFO] Leyendo archivo de Asistencia del Reloj Checador...{CTerm.END}")
        asistencia_df = pd.read_excel(ARCHIVO_ASISTENCIA, header=None)
    except Exception as e:
        print(f"{CTerm.RED}[ERROR] Error leyendo Asistencia: {e}{CTerm.END}")
        return

    mapa_dias = {}
    if len(asistencia_df) > 3:
        for col_idx, val in asistencia_df.iloc[3].items():
            try:
                if 1 <= int(float(str(val).strip())) <= 31: mapa_dias[col_idx] = int(float(str(val).strip()))
            except: continue

    nuevos_registros = []
    registro_horas_especiales = {oficial: {} for oficial in mapa_nombres_especiales.values()}
    
    # 3. PROCESAR ASISTENCIA
    print(f"{CTerm.BLUE}[INFO] Analizando tolerancias y horas de personal...{CTerm.END}")
    for i in range(len(asistencia_df) - 1):
        fila_str = [str(x).strip() for x in asistencia_df.iloc[i].values]
        idx_nombre = fila_str.index("Nombre:") if "Nombre:" in fila_str else -1
        
        if idx_nombre != -1:
            nombre_raw = next((fila_str[k] for k in range(idx_nombre + 1, len(fila_str)) if fila_str[k] not in ["", "nan"]), "")
            nombre_encontrado = normalizar_nombre(nombre_raw)
            fila_datos = asistencia_df.iloc[i+1]
            
            if nombre_encontrado in mapa_nombres_enfermeria:
                nombre_oficial = mapa_nombres_enfermeria[nombre_encontrado]
                if nombre_oficial in EXCEPCIONES_DASHBOARD: continue
                
                if nombre_oficial in TURNO_NOCHE:
                    for col_idx, dia_num in mapa_dias.items():
                        celda_hoy = str(fila_datos[col_idx]).strip()
                        if celda_hoy and celda_hoy.lower() != 'nan':
                            t_hoy = [celda_hoy[j:j+5] for j in range(0, len(celda_hoy), 5) if len(celda_hoy[j:j+5]) == 5 and ':' in celda_hoy[j:j+5]]
                            entradas = [t for t in t_hoy if time_to_mins(t) >= 14 * 60]
                            if entradas and time_to_mins(entradas[0]) >= (20 * 60 + 15):
                                nuevos_registros.append({'fecha': datetime.date(anio, mes, dia_num), 'empleado': nombre_oficial, 'incidencia': 'Retardo', 'obs': f"Entró a las {entradas[0]} (Cubrió turno noche)"})
                            col_manana = next((c for c, d in mapa_dias.items() if d == dia_num + 1), None)
                            if col_manana:
                                celda_m = str(fila_datos[col_manana]).strip()
                                if celda_m and celda_m.lower() != 'nan':
                                    t_m = [celda_m[j:j+5] for j in range(0, len(celda_m), 5) if len(celda_m[j:j+5]) == 5 and ':' in celda_m[j:j+5]]
                                    salidas = [t for t in t_m if time_to_mins(t) < 14 * 60]
                                    if salidas and time_to_mins(salidas[-1]) < (7 * 60 + 50):
                                        nuevos_registros.append({'fecha': datetime.date(anio, mes, dia_num) + datetime.timedelta(days=1), 'empleado': nombre_oficial, 'incidencia': 'Salida Anticipada', 'obs': f"Salió a las {salidas[-1]} (Cubrió turno noche)"})
                else:
                    for col_idx, dia_num in mapa_dias.items():
                        celda = str(fila_datos[col_idx]).strip()
                        if celda and celda.lower() != 'nan':
                            t = [celda[j:j+5] for j in range(0, len(celda), 5) if len(celda[j:j+5]) == 5 and ':' in celda[j:j+5]]
                            if t:
                                m_prim, m_ult = time_to_mins(t[0]), time_to_mins(t[-1])
                                if m_prim <= 14 * 60 and m_prim >= (8 * 60 + 15):
                                    nuevos_registros.append({'fecha': datetime.date(anio, mes, dia_num), 'empleado': nombre_oficial, 'incidencia': 'Retardo', 'obs': f"Entró a las {t[0]} (Cubrió turno día)"})
                                if len(t) > 1 and 14 * 60 <= m_ult < (19 * 60 + 50):
                                    nuevos_registros.append({'fecha': datetime.date(anio, mes, dia_num), 'empleado': nombre_oficial, 'incidencia': 'Salida Anticipada', 'obs': f"Salió a las {t[-1]} (Cubrió turno día)"})

            if nombre_encontrado in mapa_nombres_especiales:
                nombre_oficial = mapa_nombres_especiales[nombre_encontrado]
                for col_idx, dia_num in mapa_dias.items():
                    celda = str(fila_datos[col_idx]).strip()
                    if celda and celda.lower() != 'nan':
                        t = [celda[j:j+5] for j in range(0, len(celda), 5) if len(celda[j:j+5]) == 5 and ':' in celda[j:j+5]]
                        if len(t) >= 2:
                            m_prim, m_ult = time_to_mins(t[0]), time_to_mins(t[-1])
                            if m_ult < m_prim: m_ult += 24 * 60
                            str_semana = f"Semana {datetime.date(anio, mes, dia_num).isocalendar()[1]}"
                            registro_horas_especiales[nombre_oficial][str_semana] = registro_horas_especiales[nombre_oficial].get(str_semana, 0.0) + ((m_ult - m_prim) / 60.0)

    # 4. GUARDAR BITÁCORA EN EXCEL
    try:
        wb = load_workbook(ARCHIVO_MASTER)
        ws = wb['Auditorías'] if 'Auditorías' in wb.sheetnames else wb['BITACORA']
        if nuevos_registros:
            f_actual, u_id = encontrar_fila_inicio_y_ultimo_id(ws)
            for reg in nuevos_registros:
                u_id += 1
                ws.cell(row=f_actual, column=1, value=u_id)
                ws.cell(row=f_actual, column=2, value=reg['fecha']).number_format = 'DD/MM/YYYY'
                ws.cell(row=f_actual, column=3, value=reg['empleado'])
                ws.cell(row=f_actual, column=4, value=reg['incidencia'])
                ws.cell(row=f_actual, column=5, value=reg['obs'])
                f_actual += 1
        wb.save(ARCHIVO_MASTER)
        print(f"{CTerm.GREEN}[ÉXITO] Bitácora en Excel actualizada correctamente.{CTerm.END}")
    except Exception as e:
        print(f"{CTerm.RED}[ERROR] Guardando Excel: {e}{CTerm.END}")

    # ==========================================
    # CÁLCULO DE PAGOS Y CREACIÓN DEL PDF
    # ==========================================
    try:
        print(f"{CTerm.BLUE}[INFO] Generando gráficas y empaquetando Reporte PDF...{CTerm.END}")
        df_dash_raw = pd.read_excel(ARCHIVO_MASTER, sheet_name='DASHBOARD', header=None)
        idx_header = df_dash_raw[df_dash_raw.astype(str).apply(lambda x: x.str.contains('COLABORADOR', case=False, na=False)).any(axis=1)].index[0]
        df_dash = pd.read_excel(ARCHIVO_MASTER, sheet_name='DASHBOARD', header=idx_header)
        df_dash = df_dash.dropna(subset=['COLABORADOR'])
        
        df_bitacora = pd.read_excel(ARCHIVO_MASTER, sheet_name='Auditorías' if 'Auditorías' in load_workbook(ARCHIVO_MASTER).sheetnames else 'BITACORA')
        
        if 'FECHA' in df_bitacora.columns:
            df_bitacora['FECHA'] = pd.to_datetime(df_bitacora['FECHA'], errors='coerce')
            df_bitacora = df_bitacora[(df_bitacora['FECHA'].dt.year == anio) & (df_bitacora['FECHA'].dt.month == mes)]

        pagos = {}
        for _, row in df_dash.iterrows():
            colaborador = str(row['COLABORADOR']).strip()
            if colaborador.lower() == 'nan' or colaborador == '': continue
            if colaborador in EXCEPCIONES_DASHBOARD: continue
            
            pagos[colaborador] = {
                'Puntualidad': 500, 'Uniforme': 500, 'Admin': 500, 'Total': 1500,
                'Cant_Retardos': 0, 'Fechas_Retardos': [], 'Falta_Uniforme': False, 'Falta_Admin': False, 'Agresion': False,
                'Penalizaciones': set(), 'Obs_Admin': ""
            }

        conteo_fallas = {"Regla de Oro": 0, "Más de 3 retardos acumulados": 0, "Uniforme": 0, "Kaizen faltante/Capacitacion": 0}
        retardos_totales_grafica = {}
        historial_persona = {}

        for _, row in df_bitacora.iterrows():
            emp = str(row['EMPLEADO']).strip()
            inc = str(row['INCIDENCIA']).strip().lower()
            obs = str(row['OBSERVACIÓN']).strip() if 'OBSERVACIÓN' in df_bitacora.columns else ""
            fecha_str = row['FECHA'].strftime('%d/%m/%Y') if pd.notna(row['FECHA']) else ""
            
            if emp and emp.lower() != 'nan':
                if emp not in historial_persona: historial_persona[emp] = []
                historial_persona[emp].append({'fecha': fecha_str, 'incidencia': str(row['INCIDENCIA']).strip(), 'obs': obs})

            if emp in pagos:
                if "retardo" in inc:
                    pagos[emp]['Cant_Retardos'] += 1
                    pagos[emp]['Fechas_Retardos'].append(fecha_str)
                    retardos_totales_grafica[emp] = retardos_totales_grafica.get(emp, 0) + 1
                elif "uniforme" in inc or "uniforme" in obs.lower():
                    pagos[emp]['Falta_Uniforme'] = True
                elif "admin" in inc or "kaizen" in inc or "capacitaci" in inc or "admin" in obs.lower() or "kaizen" in obs.lower() or "capacitaci" in obs.lower():
                    pagos[emp]['Falta_Admin'] = True
                    pagos[emp]['Obs_Admin'] = obs
                elif "agresi" in inc or "conflicto" in inc or "agresi" in obs.lower():
                    pagos[emp]['Agresion'] = True

        for emp, p in pagos.items():
            fechas_texto = f" ({', '.join(p['Fechas_Retardos'])})" if p['Fechas_Retardos'] else ""

            if p['Cant_Retardos'] == 2:
                p['Penalizaciones'].add(f"[ALERTA] Acumulo 2 retardos{fechas_texto}. Estuvo a un retardo de perder el bono. Favor de hablar con esta persona.")
                
            if p['Cant_Retardos'] >= 3 and p['Cant_Retardos'] < 6:
                p['Puntualidad'] = 0
                p['Penalizaciones'].add(f"Bono Perdido: Acumulo {p['Cant_Retardos']} Retardos{fechas_texto}")
                conteo_fallas["Más de 3 retardos acumulados"] += 1
            
            if p['Falta_Uniforme']:
                p['Uniforme'] = 0
                p['Penalizaciones'].add("Bono Perdido: Falta de Uniforme")
                conteo_fallas["Uniforme"] += 1

            if p['Falta_Admin']:
                p['Admin'] = 0
                motivo = f"Bono Perdido: {p['Obs_Admin']}" if p['Obs_Admin'] else "Bono Perdido: Falta de Kaizen o Capacitacion"
                p['Penalizaciones'].add(motivo)
                conteo_fallas["Kaizen faltante/Capacitacion"] += 1

            if p['Cant_Retardos'] >= 6 or p['Agresion']:
                p['Puntualidad'] = 0; p['Uniforme'] = 0; p['Admin'] = 0
                if p['Cant_Retardos'] >= 6:
                    p['Penalizaciones'].add(f"REGLA DE ORO ROTA: Acumulo {p['Cant_Retardos']} Retardos totales{fechas_texto}")
                if p['Agresion']:
                    p['Penalizaciones'].add("REGLA DE ORO ROTA: Agresion o Conflicto")
                conteo_fallas["Regla de Oro"] += 1

            p['Total'] = p['Puntualidad'] + p['Uniforme'] + p['Admin']
            p['Penalizaciones'] = list(p['Penalizaciones'])

        if sum(conteo_fallas.values()) > 0:
            plt.figure(figsize=(6, 3.5))
            ax = plt.gca()
            nombres = [k for k, v in conteo_fallas.items() if v > 0]
            valores = [v for v in conteo_fallas.values() if v > 0]
            nombres = [x for _, x in sorted(zip(valores, nombres), reverse=True)]
            valores = sorted(valores, reverse=True)
            
            barras = ax.bar(nombres, valores, color='#1F3A52', width=0.4) 
            ax.spines['top'].set_visible(False); ax.spines['right'].set_visible(False)
            ax.yaxis.set_major_locator(MaxNLocator(integer=True))
            ax.yaxis.grid(True, linestyle='--', alpha=0.5)
            ax.set_axisbelow(True)
            plt.xticks(fontsize=8)
            for bar in barras:
                ax.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 0.1, int(bar.get_height()), ha='center', va='bottom', fontsize=9, fontweight='bold', color='#2C3E50')
            plt.tight_layout()
            plt.savefig('grafico_pareto.png', transparent=True, dpi=300)
            plt.close()

        if retardos_totales_grafica:
            plt.figure(figsize=(5.5, 3))
            ax = plt.gca()
            ordenados = sorted(retardos_totales_grafica.items(), key=lambda x: x[1])
            nombres_h = [abreviar_nombre(x[0]) for x in ordenados]
            valores_h = [x[1] for x in ordenados]
            
            barras_h = ax.barh(nombres_h, valores_h, color='#D35400', height=0.6)
            ax.spines['top'].set_visible(False); ax.spines['right'].set_visible(False); ax.spines['bottom'].set_visible(False)
            ax.xaxis.set_major_locator(MaxNLocator(integer=True))
            ax.xaxis.grid(True, linestyle='--', alpha=0.5)
            ax.set_axisbelow(True)
            for bar in barras_h:
                ax.text(bar.get_width() + 0.1, bar.get_y() + bar.get_height()/2, int(bar.get_width()), ha='left', va='center', fontsize=9, fontweight='bold', color='#2C3E50')
            plt.tight_layout()
            plt.savefig('grafico_retardos.png', transparent=True, dpi=300)
            plt.close()

        pdf = PremiumPDF()
        pdf.set_auto_page_break(auto=True, margin=15)
        pdf.cover_page(nombre_mes_actual, anio)
        pdf.add_page()
        
        pdf.chapter_title("1. Análisis de Incidencias (Pareto)")
        pdf.set_font('Helvetica', '', 10)
        
        if sum(conteo_fallas.values()) > 0:
            pdf.multi_cell(0, 5, "Motivos principales de pérdida de bono en el periodo actual:")
            if os.path.exists('grafico_pareto.png'):
                pdf.image('grafico_pareto.png', x=30, w=150)
                os.remove('grafico_pareto.png')
        else:
            pdf.multi_cell(0, 6, "Excelente: No se detectaron penalizaciones en este periodo.")
            
        pdf.ln(2)

        pdf.chapter_title("Ranking de Retardos")
        pdf.set_font('Helvetica', '', 10)
        if retardos_totales_grafica:
            pdf.multi_cell(0, 5, "Historial de retardos acumulados durante el periodo.")
            if os.path.exists('grafico_retardos.png'):
                pdf.image('grafico_retardos.png', x=30, w=150)
                os.remove('grafico_retardos.png')
        else:
            pdf.multi_cell(0, 6, "Excelente: Cero retardos registrados en todo el equipo.")

        pdf.add_page()
        pdf.chapter_title("2. TOTAL A PAGAR")
        pdf.set_font('Helvetica', 'B', 9)
        pdf.set_fill_color(*C_NAVY)
        pdf.set_text_color(255, 255, 255)
        
        w_col, w_num = 66, 30
        ancho_total = w_col + (w_num * 4)
        margen_izq = (210 - ancho_total) / 2
        
        pdf.set_x(margen_izq)
        pdf.cell(w_col, 8, 'COLABORADOR', 1, 0, 'C', True)
        pdf.cell(w_num, 8, 'PUNTUALIDAD', 1, 0, 'C', True)
        pdf.cell(w_num, 8, 'UNIFORME', 1, 0, 'C', True)
        pdf.cell(w_num, 8, 'ADMIN', 1, 0, 'C', True)
        pdf.cell(w_num, 8, 'TOTAL', 1, 1, 'C', True)

        pdf.set_font('Helvetica', '', 9)
        total_empresa = 0
        fill_row = False
        pagos_ordenados = dict(sorted(pagos.items()))
        
        for emp, b in pagos_ordenados.items():
            pdf.set_x(margen_izq)
            pdf.set_fill_color(*C_LIGHT_BG) if fill_row else pdf.set_fill_color(255, 255, 255)
            
            pdf.set_text_color(*C_DARK)
            pdf.cell(w_col, 8, emp, 1, 0, 'C', fill_row)
            
            pdf.set_text_color(*C_SUN) if b['Puntualidad'] == 0 else pdf.set_text_color(*C_DARK)
            pdf.cell(w_num, 8, f"${b['Puntualidad']}", 1, 0, 'C', fill_row)
            
            pdf.set_text_color(*C_SUN) if b['Uniforme'] == 0 else pdf.set_text_color(*C_DARK)
            pdf.cell(w_num, 8, f"${b['Uniforme']}", 1, 0, 'C', fill_row)
            
            pdf.set_text_color(*C_SUN) if b['Admin'] == 0 else pdf.set_text_color(*C_DARK)
            pdf.cell(w_num, 8, f"${b['Admin']}", 1, 0, 'C', fill_row)
            
            pdf.set_text_color(*C_DARK)
            pdf.set_font('Helvetica', 'B', 9)
            pdf.cell(w_num, 8, f"${b['Total']}", 1, 1, 'C', fill_row)
            pdf.set_font('Helvetica', '', 9)
            
            total_empresa += b['Total']
            fill_row = not fill_row

        pdf.ln(4)
        pdf.set_font('Helvetica', 'B', 12)
        pdf.set_text_color(*C_NAVY)
        pdf.cell(0, 8, f"GRAN TOTAL A DISPERSAR: ${total_empresa} MXN", 0, 1, 'C')

        pdf.add_page()
        pdf.chapter_title("3. Justificacion de Retenciones y Advertencias")
        pdf.set_font('Helvetica', '', 10)
        pdf.set_text_color(50, 50, 50)
        
        hubo_retenidos = False
        for emp, b in pagos_ordenados.items():
            if b['Penalizaciones']:
                hubo_retenidos = True
                pdf.set_font('Helvetica', 'B', 11)
                pdf.set_text_color(*C_NAVY)
                pdf.cell(0, 6, f"- {emp} (Total a pagar: ${b['Total']})", 0, 1) 
                
                pdf.set_font('Helvetica', '', 10)
                pdf.set_text_color(*C_DARK)
                pdf.cell(10)
                pdf.multi_cell(0, 6, f"{', '.join(b['Penalizaciones'])}")
                pdf.ln(3)
                
        if not hubo_retenidos:
            pdf.cell(0, 6, "Excelente periodo. Todos mantienen su bono al 100% y no hay advertencias.", 0, 1)

        if registro_horas_especiales:
            pdf.add_page()
            pdf.chapter_title("4. Control de Horas Semanales")
            pdf.set_font('Helvetica', 'B', 9)
            pdf.set_fill_color(*C_NAVY)
            pdf.set_text_color(255, 255, 255)
            
            w_col_h, w_sem, w_hrs = 70, 58, 58
            ancho_total_h = w_col_h + w_sem + w_hrs
            margen_izq_h = (210 - ancho_total_h) / 2
            
            pdf.set_x(margen_izq_h)
            pdf.cell(w_col_h, 8, 'COLABORADOR', 1, 0, 'C', True)
            pdf.cell(w_sem, 8, 'SEMANA', 1, 0, 'C', True)
            pdf.cell(w_hrs, 8, 'HORAS TRABAJADAS', 1, 1, 'C', True)

            pdf.set_font('Helvetica', '', 9)
            fill_row_h = False
            for emp, sems in registro_horas_especiales.items():
                if not sems: continue
                pdf.set_text_color(*C_DARK)
                pdf.set_x(margen_izq_h)
                pdf.set_fill_color(*C_LIGHT_BG) if fill_row_h else pdf.set_fill_color(255, 255, 255)
                
                pdf.set_font('Helvetica', 'B', 9)
                pdf.cell(w_col_h, 8 * len(sems), emp, 1, 0, 'C', fill_row_h)
                pdf.set_font('Helvetica', '', 9)
                
                x_act, y_act = pdf.get_x(), pdf.get_y()
                for i, (sem, hrs) in enumerate(sems.items()):
                    pdf.set_xy(x_act, y_act + (i * 8))
                    pdf.cell(w_sem, 8, sem, 1, 0, 'C', fill_row_h)
                    
                    pdf.set_text_color(*C_SUN) if hrs < 40 else pdf.set_text_color(*C_NAVY)
                    pdf.set_font('Helvetica', 'B', 9)
                    pdf.cell(w_hrs, 8, f"{round(hrs, 1)} hrs", 1, 0, 'C', fill_row_h)
                    pdf.set_font('Helvetica', '', 9)
                    pdf.set_text_color(*C_DARK)
                pdf.ln(8)
                fill_row_h = not fill_row_h

        pdf.add_page()
        pdf.chapter_title("5. Bitacora de Auditoria")
        
        if historial_persona:
            for emp, registros in historial_persona.items():
                if emp not in pagos_ordenados: continue
                
                if pdf.get_y() > 235:
                    pdf.add_page()
                
                pdf.set_font('Helvetica', 'B', 11)
                pdf.set_text_color(*C_NAVY)
                pdf.cell(0, 8, f"Colaborador(a): {emp}", 0, 1, 'L')
                
                pdf.set_font('Helvetica', 'B', 9)
                pdf.set_fill_color(*C_NAVY)
                pdf.set_text_color(255, 255, 255)
                
                w_f, w_i, w_o = 30, 50, 110  
                margen_izq_bit = (210 - (w_f + w_i + w_o)) / 2
                
                pdf.set_x(margen_izq_bit)
                pdf.cell(w_f, 7, 'FECHA', 1, 0, 'C', True)
                pdf.cell(w_i, 7, 'INCIDENCIA', 1, 0, 'C', True)
                pdf.cell(w_o, 7, 'OBSERVACIÓN', 1, 1, 'C', True)

                pdf.set_font('Helvetica', '', 8)
                fill_row = False
                
                for reg in registros:
                    if pdf.get_y() > 260:
                        pdf.add_page()
                        pdf.set_x(margen_izq_bit)
                        pdf.set_font('Helvetica', 'B', 9)
                        pdf.set_fill_color(*C_NAVY)
                        pdf.set_text_color(255, 255, 255)
                        pdf.cell(w_f, 7, 'FECHA', 1, 0, 'C', True)
                        pdf.cell(w_i, 7, 'INCIDENCIA', 1, 0, 'C', True)
                        pdf.cell(w_o, 7, 'OBSERVACIÓN', 1, 1, 'C', True)
                        pdf.set_font('Helvetica', '', 8)

                    pdf.set_x(margen_izq_bit)
                    pdf.set_fill_color(*C_LIGHT_BG) if fill_row else pdf.set_fill_color(255, 255, 255)
                    pdf.set_text_color(*C_DARK)
                    
                    pdf.cell(w_f, 6, reg['fecha'], 1, 0, 'C', fill_row)
                    
                    pdf.set_text_color(*C_SUN)
                    pdf.cell(w_i, 6, reg['incidencia'], 1, 0, 'C', fill_row)
                    
                    pdf.set_text_color(*C_DARK)
                    obs_text = reg['obs']
                    pdf.cell(w_o, 6, obs_text, 1, 1, 'L', fill_row)
                    
                    fill_row = not fill_row
                
                pdf.ln(6) 
        else:
            pdf.set_font('Helvetica', '', 9)
            pdf.cell(0, 6, "No hay historial disponible para este periodo.", 0, 1)

        # --- LÓGICA DE CARPETAS INTELIGENTES ---
        carpeta_historial = os.path.join(CARPETA_ACTUAL, "Historial_Reportes")
        nombre_carpeta_mes = f"{mes:02d}_{nombre_mes_actual}_{anio}"
        ruta_carpeta_mes = os.path.join(carpeta_historial, nombre_carpeta_mes)
        
        os.makedirs(ruta_carpeta_mes, exist_ok=True)

        nombre_pdf = f"Reporte_Sunhaven_{mes:02d}_{anio}.pdf"
        ruta_guardado = os.path.join(ruta_carpeta_mes, nombre_pdf)
        pdf.output(ruta_guardado)
        
        print(f"\n{CTerm.GREEN}{CTerm.BOLD}[ÉXITO TOTAL] Reporte guardado en: Historial_Reportes/{nombre_carpeta_mes}/{nombre_pdf}{CTerm.END}\n")

    except Exception as e:
        print(f"{CTerm.RED}[ERROR] En el calculo o generacion de PDF: {e}{CTerm.END}")

if __name__ == "__main__":
    main()