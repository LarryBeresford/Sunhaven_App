import streamlit as st
import pandas as pd
import gspread
import json
import os
import io
from datetime import datetime, timedelta
import plotly.express as px
from fpdf import FPDF
from openpyxl import load_workbook
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import shutil
import uuid

# ==========================================
# 0. CONFIGURACION Y SEGURIDAD GLOBAL
# ==========================================
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
PATH_CREDS = os.path.join(BASE_DIR, 'config', 'sunhaven-credentials.json')
PATH_BITACORA = os.path.join(BASE_DIR, 'data', 'bitacora_interna.csv')

st.set_page_config(page_title="Sunhaven Intelligence Suite", layout="wide", initial_sidebar_state="expanded")

st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;600;800&display=swap');
    html, body, [class*="css"] { font-family: 'Inter', sans-serif; }
    .block-container { padding-top: 1.5rem; max-width: 95%; }
    h1, h2, h3 { color: #1e293b; font-weight: 800; letter-spacing: -0.5px; }
    .exec-header { background-color: #ffffff; padding: 2rem; border-radius: 4px; border: 1px solid #e2e8f0; box-shadow: 0 1px 3px rgba(0,0,0,0.1); margin-bottom: 2rem; display: flex; justify-content: space-between; align-items: center; }
    .kpi-card { background-color: #ffffff; padding: 25px; border-radius: 4px; border: 1px solid #e2e8f0; text-align: center; }
    .metric-label { font-size: 12px; color: #64748b; font-weight: 700; margin-bottom: 8px; text-transform: uppercase; letter-spacing: 1px; }
    .metric-value { font-size: 34px; font-weight: 800; margin: 0; }
    .footer-watermark { text-align: center; margin-top: 80px; opacity: 0.3; font-size: 11px; font-weight: 600; letter-spacing: 2px; text-transform: uppercase; }
    .stTabs [data-baseweb="tab-list"] { gap: 10px; }
    .stTabs [data-baseweb="tab"] { height: 45px; background-color: #f1f5f9; border-radius: 4px; padding: 10px 20px; font-weight: 600; color: #475569; }
    .stTabs [aria-selected="true"] { background-color: #1e293b !important; color: white !important; }
    #MainMenu {visibility: hidden;}
    footer {visibility: hidden;}
</style>
""", unsafe_allow_html=True)

HEX_NAVY, HEX_RED, HEX_GREEN, HEX_SUN = "#1e293b", "#dc2626", "#16a34a", "#d35400"
C_NAVY, C_SUN, C_DARK, C_LIGHT = (31, 58, 82), (211, 84, 0), (44, 62, 80), (245, 247, 248)

# --- REGLAS DE NEGOCIO ---
ENFERMERAS_ROL_A = ["Consuelo Ceja Liborio", "Jaqueline Hernández Sosa"]
ENFERMERAS_ROL_B = ["Silvia Rodríguez Reynaga", "Guadalupe Georgia Lopez Ceja"]
ENFERMERAS_NOCHE = ENFERMERAS_ROL_A + ENFERMERAS_ROL_B

EMPLEADOS_DB = {
    "ALE": "Alejandra Itzel de la Fuente Ramírez", "ARIANA": "Ariana Villanueva Temores",
    "Araceli Figueroa": "Blanca Aracely Figueroa Marroquín", "BLANCARUVC": "Blanca Estela Ruvalcaba Ruiz",
    "CARMEN": "Carmen Torres Reyes", "ConsueloCeja": "Consuelo Ceja Liborio", "CRISTY": "Cristina Ramos Aquino", 
    "HUGO": "Hugo Silva Esparza", "JACK": "Jaqueline Hernández Sosa", "CESAR": "Julio César Pérez Carranza",
    "MARIASEO": "Maribel Herrera Mauricio", "MarthaCastro": "Martha Manuela Castro García",
    "MAYT": "Mayte López Romero", "MONI": "Mónica Teresa Grande Figueroa", "NANCI": "Nancy Estephania González Velasco", 
    "NIRE": "Nireida Flores Núñez", "Olga Gabriela Jimenez": "Olga Gabriela Jiménez Medina", 
    "RosaCastro": "Rosa Isela Antonieta Castro García", "SANDY": "Sandy Yusbeth Cruz González", 
    "Silvia Rodriguez": "Silvia Rodríguez Reynaga", "VERO": "Verónica Janeth Gómez López", 
    "YamileLuquin": "Yareli Yamile Luquin Puga", "Guadalupe": "Guadalupe Georgia Lopez Ceja"
}

ENFERMERAS_LISTA = [
    "Ariana Villanueva Temores", "Blanca Aracely Figueroa Marroquín", "Blanca Estela Ruvalcaba Ruiz",
    "Consuelo Ceja Liborio", "Cristina Ramos Aquino", "Jaqueline Hernández Sosa", "Mayte López Romero",
    "Nancy Estephania González Velasco", "Nireida Flores Núñez", "Olga Gabriela Jiménez Medina",
    "Rosa Isela Antonieta Castro García", "Sandy Yusbeth Cruz González", "Silvia Rodríguez Reynaga",
    "Verónica Janeth Gómez López", "Yareli Yamile Luquin Puga", "Guadalupe Georgia Lopez Ceja"
]

EXCEPCIONES_KAIZEN = ["Yareli Yamile Luquin Puga", "Blanca Aracely Figueroa Marroquín", "Guadalupe Georgia Lopez Ceja"]
HORA_ENTRADA_DIA, HORA_ENTRADA_NOCHE = datetime.strptime("08:15", "%H:%M").time(), datetime.strptime("20:15", "%H:%M").time()
TIPO_INCIDENCIAS = ["Falta de uniforme (Leve)", "Uso de celular (Leve)", "No hacer entrega (Leve)", "No hacer ronda (Leve)", "Salida anticipada (Leve)", "AGRESIÓN / CONFLICTO (Grave)", "REGLA DE ORO (Grave)"]

# ==========================================
# 1. CLASE MAESTRA DE PDF Y HELPERS
# ==========================================
def sanitizar_texto(texto):
    return str(texto).replace('•', '-').replace('“', '"').replace('”', '"').replace('–', '-').encode('latin-1', 'replace').decode('latin-1')

def tabla_centrada(pdf, headers, data, col_widths):
    start_x = (210 - sum(col_widths)) / 2
    pdf.set_x(start_x)
    pdf.set_fill_color(*C_NAVY)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font('Helvetica', 'B', 8)
    for i, h in enumerate(headers): pdf.cell(col_widths[i], 7, sanitizar_texto(h), 1, 0, 'C', True)
    pdf.ln()
    pdf.set_font('Helvetica', '', 8)
    fill = False
    for row in data:
        pdf.set_x(start_x)
        pdf.set_fill_color(*C_LIGHT) if fill else pdf.set_fill_color(255, 255, 255)
        pdf.set_text_color(*C_DARK)
        for i, val in enumerate(row):
            align = 'C' if i > 0 else 'L'
            pdf.cell(col_widths[i], 6, sanitizar_texto(str(val)), 1, 0, align, fill)
        pdf.ln()
        fill = not fill

class SunhavenPDF(FPDF):
    def header(self):
        if self.page_no() == 1: return
        self.set_fill_color(*C_NAVY)
        self.rect(0, 0, 210, 26, 'F')
        self.set_fill_color(*C_SUN)
        self.rect(0, 26, 210, 1.5, 'F')
        self.set_y(9)
        self.set_font('Helvetica', 'B', 16)
        self.set_text_color(255, 255, 255)
        self.cell(0, 8, sanitizar_texto(getattr(self, 'titulo_header', 'REPORTE EJECUTIVO')), 0, 1, 'C')
        self.ln(12)

    def footer(self):
        if self.page_no() == 1: return
        self.set_y(-15)
        self.set_font('Helvetica', 'I', 8)
        self.set_text_color(*C_NAVY)
        self.set_draw_color(*C_SUN)
        self.line(10, 282, 200, 282)
        self.set_x(10)
        self.cell(190, 10, sanitizar_texto(f'Página {self.page_no()}'), 0, 0, 'C')
        self.set_x(10)
        self.cell(190, 10, sanitizar_texto('Ing. Larry Beresford'), 0, 0, 'R')

    def cover_page(self, titulo_principal, subtitulo, fecha_str):
        self.add_page()
        self.set_fill_color(*C_NAVY)
        self.rect(0, 0, 210, 100, 'F')
        self.set_fill_color(*C_SUN)
        self.rect(0, 100, 210, 3, 'F')
        self.set_y(40)
        self.set_font('Helvetica', 'B', 28)
        self.set_text_color(255, 255, 255)
        self.cell(0, 10, sanitizar_texto('SUN HAVEN'), 0, 1, 'C')
        self.set_font('Helvetica', '', 12)
        self.set_text_color(220, 220, 220)
        self.cell(0, 8, sanitizar_texto('CASA DE DESCANSO PARA ADULTOS MAYORES'), 0, 1, 'C')
        self.set_y(140)
        self.set_font('Helvetica', 'B', 22)
        self.set_text_color(*C_DARK)
        self.cell(0, 10, sanitizar_texto('REPORTE EJECUTIVO'), 0, 1, 'C')
        self.set_font('Helvetica', 'B', 22)
        self.set_text_color(*C_SUN)
        self.cell(0, 12, sanitizar_texto(titulo_principal), 0, 1, 'C')
        self.set_font('Helvetica', '', 14)
        self.set_text_color(*C_DARK)
        self.cell(0, 8, sanitizar_texto(subtitulo), 0, 1, 'C')
        self.set_y(220)
        self.set_font('Helvetica', 'B', 12)
        self.set_text_color(*C_NAVY)
        self.cell(0, 8, sanitizar_texto(f'PERIODO EVALUADO: {fecha_str}'), 0, 1, 'C')
        self.set_y(260)
        self.set_font('Helvetica', '', 11)
        self.set_text_color(*C_DARK)
        self.cell(0, 6, sanitizar_texto('Elaborado por:'), 0, 1, 'C')
        self.set_font('Helvetica', 'B', 11)
        self.cell(0, 6, sanitizar_texto('Ing. Larry Beresford'), 0, 1, 'C')

# ==========================================
# 2. GENERADORES DE PDF POR MÓDULO
# ==========================================
def generar_pdf_dashboard_op(ico, estatus, df_a, df_c, df_evol, fecha_str):
    temp_dir = os.path.join(os.path.dirname(__file__), f'temp_img_{uuid.uuid4().hex}')
    os.makedirs(temp_dir, exist_ok=True)
    
    plt.figure(figsize=(7, 4))
    plt.bar(df_a['index'], df_a['V'], color=HEX_NAVY)
    plt.axhline(90, color='red', linestyle='--')
    plt.title('Pareto por Área Operativa', fontsize=10, fontweight='bold')
    plt.tight_layout()
    p_pareto = os.path.join(temp_dir, 'pareto_op.png')
    plt.savefig(p_pareto)
    plt.close()

    plt.figure(figsize=(7, 4))
    df_c_sorted = df_c.sort_values('V', ascending=True)
    plt.barh(df_c_sorted['index'], df_c_sorted['V'], color=HEX_RED)
    plt.axvline(90, color='black', linestyle='--')
    plt.title('Análisis de Causa Raíz', fontsize=10, fontweight='bold')
    plt.tight_layout()
    p_causa = os.path.join(temp_dir, 'causa_op.png')
    plt.savefig(p_causa)
    plt.close()
    
    plt.figure(figsize=(9, 4))
    for col in df_evol.columns:
        plt.plot(df_evol.index, df_evol[col], marker='o', label=col)
    plt.axhline(90, color='red', linestyle='--')
    plt.title('Evolución Histórica', fontsize=10, fontweight='bold')
    plt.legend(loc='lower center', bbox_to_anchor=(0.5, -0.3), ncol=5)
    plt.tight_layout()
    p_evol = os.path.join(temp_dir, 'evol_op.png')
    plt.savefig(p_evol)
    plt.close()

    pdf = SunhavenPDF()
    pdf.titulo_header = "REPORTE EJECUTIVO - OPERACIONES"
    pdf.cover_page("INDICADORES OPERATIVOS (KPI)", "Estado de las Infraestructuras y Servicios", fecha_str)
    
    pdf.add_page()
    pdf.set_font('Helvetica', 'B', 14)
    pdf.set_text_color(*C_NAVY)
    pdf.cell(0, 8, sanitizar_texto("1. ESTATUS GLOBAL INSTITUCIONAL"), 0, 1, 'L')
    pdf.set_draw_color(*C_SUN)
    pdf.line(10, pdf.get_y(), 200, pdf.get_y())
    pdf.ln(5)
    pdf.set_font('Helvetica', '', 11)
    pdf.set_text_color(*C_DARK)
    pdf.cell(0, 8, sanitizar_texto(f"Índice de Cumplimiento Operativo (ICO Maestro): {ico:.1f}%"), 0, 1)
    pdf.cell(0, 8, sanitizar_texto(f"Dictamen del Sistema: {estatus}"), 0, 1)
    pdf.ln(5)
    
    areas_bajas = df_a[df_a['V'] < 90].sort_values('V', ascending=True)
    pdf.set_font('Helvetica', 'B', 10)
    pdf.set_text_color(*C_SUN)
    if not areas_bajas.empty:
        pdf.multi_cell(0, 6, sanitizar_texto("DICTAMEN: Las siguientes áreas presentan un desempeño inferior al 90% y requieren atención por orden de prioridad:"))
        pdf.set_font('Helvetica', '', 10)
        for _, r in areas_bajas.iterrows():
            pdf.cell(0, 6, sanitizar_texto(f"  - {r['index']}: {r['V']:.1f}%"), 0, 1)
    else:
        pdf.multi_cell(0, 6, sanitizar_texto("DICTAMEN: Todas las áreas operan de manera óptima por encima de la línea base del 90%."))
    pdf.ln(5)

    pdf.set_font('Helvetica', 'B', 14)
    pdf.set_text_color(*C_NAVY)
    pdf.cell(0, 8, sanitizar_texto("2. DESEMPEÑO POR DEPARTAMENTO"), 0, 1, 'L')
    pdf.line(10, pdf.get_y(), 200, pdf.get_y())
    pdf.image(p_pareto, x=30, w=150)
    pdf.ln(80)
    
    pdf.add_page()
    pdf.set_font('Helvetica', 'B', 14)
    pdf.set_text_color(*C_NAVY)
    pdf.cell(0, 8, sanitizar_texto("3. ANÁLISIS DE CAUSA RAÍZ"), 0, 1, 'L')
    pdf.line(10, pdf.get_y(), 200, pdf.get_y())
    pdf.ln(3)

    causas_bajas = df_c[df_c['V'] < 90].sort_values('V', ascending=True)
    pdf.set_font('Helvetica', 'B', 10)
    pdf.set_text_color(*C_SUN)
    if not causas_bajas.empty:
        pdf.multi_cell(0, 6, sanitizar_texto("DICTAMEN: Para estabilizar el ICO, se debe atender prioritariamente la corrección de los siguientes criterios (Por debajo del 90%):"))
        pdf.set_font('Helvetica', '', 10)
        for _, r in causas_bajas.iterrows():
            pdf.cell(0, 6, sanitizar_texto(f"  - {r['index']}: {r['V']:.1f}%"), 0, 1)
    else:
        pdf.multi_cell(0, 6, sanitizar_texto("DICTAMEN: No se detectan criterios críticos en este periodo."))
    
    pdf.image(p_causa, x=30, w=150)
    pdf.ln(80)
    
    pdf.add_page()
    pdf.set_font('Helvetica', 'B', 14)
    pdf.set_text_color(*C_NAVY)
    pdf.cell(0, 8, sanitizar_texto("4. EVOLUCIÓN HISTÓRICA (TENDENCIA)"), 0, 1, 'L')
    pdf.line(10, pdf.get_y(), 200, pdf.get_y())
    pdf.ln(5)
    pdf.image(p_evol, x=15, w=180)
    
    shutil.rmtree(temp_dir, ignore_errors=True)
    return pdf.output(dest='S').encode('latin-1', 'replace')

def generar_pdf_nomina(df_nomina, df_incidencias, df_retardos, stats_kaizen, propuestas, mes_str):
    temp_dir = os.path.join(os.path.dirname(__file__), f'temp_img_{uuid.uuid4().hex}')
    os.makedirs(temp_dir, exist_ok=True)
    
    plt.figure(figsize=(7, 4))
    inc_counts = df_incidencias['INCIDENCIA'].value_counts()
    if not inc_counts.empty:
        plt.bar(inc_counts.index, inc_counts.values, color=HEX_NAVY)
        plt.title('Pareto de Incidencias', fontsize=10, fontweight='bold')
        plt.xticks(rotation=15, ha='right', fontsize=8)
    plt.tight_layout()
    p_pareto = os.path.join(temp_dir, 'pareto_nom.png')
    plt.savefig(p_pareto)
    plt.close()

    plt.figure(figsize=(7, 4))
    ret_counts = df_retardos['EMPLEADO'].value_counts().head(10)
    if not ret_counts.empty:
        plt.barh(ret_counts.index, ret_counts.values, color=HEX_RED)
        plt.title('Ranking de Retardos Acumulados', fontsize=10, fontweight='bold')
        plt.gca().invert_yaxis()
    plt.tight_layout()
    p_ret = os.path.join(temp_dir, 'ret_nom.png')
    plt.savefig(p_ret)
    plt.close()

    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(8, 4))
    colors = [HEX_GREEN, HEX_RED]
    if (stats_kaizen['prev_si'] + stats_kaizen['prev_no']) > 0: ax1.pie([stats_kaizen['prev_si'], stats_kaizen['prev_no']], labels=['Sí', 'No'], autopct='%1.1f%%', colors=colors, startangle=90)
    ax1.set_title('Participación Mes Anterior', fontsize=10)
    if (stats_kaizen['curr_si'] + stats_kaizen['curr_no']) > 0: ax2.pie([stats_kaizen['curr_si'], stats_kaizen['curr_no']], labels=['Sí', 'No'], autopct='%1.1f%%', colors=colors, startangle=90)
    ax2.set_title('Participación Mes Actual', fontsize=10)
    plt.tight_layout()
    p_kz = os.path.join(temp_dir, 'kz_nom.png')
    plt.savefig(p_kz)
    plt.close()

    pdf = SunhavenPDF()
    pdf.titulo_header = "REPORTE EJECUTIVO DE BONOS E INCIDENCIAS"
    pdf.cover_page("BONOS", "Periodo Evaluado", mes_str)
    
    pdf.add_page()
    pdf.set_font('Helvetica', 'B', 14)
    pdf.set_text_color(*C_NAVY)
    pdf.cell(0, 8, sanitizar_texto("1. ANÁLISIS DE INCIDENCIAS (PARETO)"), 0, 1, 'L')
    pdf.image(p_pareto, x=10, w=90)
    pdf.image(p_ret, x=110, w=90)
    pdf.ln(80)
    
    pdf.add_page()
    pdf.set_font('Helvetica', 'B', 14)
    pdf.set_text_color(*C_NAVY)
    pdf.cell(0, 8, sanitizar_texto("2. TOTAL A PAGAR"), 0, 1, 'L')
    datos_nomina = [[row['COLABORADOR'], f"${row['$ PUNTUAL']}", f"${row['$ UNIFORM']}", f"${row['$ ADMIN']}", f"${row['TOTAL A PAGAR']}"] for _, row in df_nomina.iterrows()]
    tabla_centrada(pdf, ["COLABORADOR", "PUNTUALIDAD", "UNIFORME", "ADMIN", "TOTAL"], datos_nomina, [70, 25, 25, 25, 25])
    
    pdf.add_page()
    pdf.titulo_header = "REPORTE EJECUTIVO DE MEJORA CONTINUA (KAIZEN)"
    pdf.set_font('Helvetica', 'B', 14)
    pdf.set_text_color(*C_NAVY)
    pdf.cell(0, 8, sanitizar_texto("1. ESTADÍSTICAS DE PARTICIPACIÓN Y TENDENCIA"), 0, 1, 'L')
    pdf.image(p_kz, x=20, w=170)
    pdf.ln(85)
    pdf.set_font('Helvetica', 'B', 10)
    pdf.set_text_color(*C_DARK)
    pdf.cell(0, 6, sanitizar_texto("Colaboradores que NO participaron:"), 0, 1)
    pdf.set_font('Helvetica', '', 10)
    for emp in stats_kaizen['lista_no']: pdf.cell(0, 5, sanitizar_texto(f"  - {emp}"), 0, 1)
        
    pdf.add_page()
    pdf.set_font('Helvetica', 'B', 14)
    pdf.set_text_color(*C_NAVY)
    pdf.cell(0, 8, sanitizar_texto("2. PROPUESTAS DE MEJORA"), 0, 1, 'L')
    for prop in propuestas:
        pdf.set_fill_color(240, 245, 250)
        pdf.set_font('Helvetica', 'B', 10)
        pdf.set_text_color(*C_NAVY)
        pdf.cell(130, 8, sanitizar_texto(f" Colaborador(a): {prop['nombre']}"), 'L T', 0, 'L', True)
        pdf.set_font('Helvetica', 'I', 9)
        pdf.cell(60, 8, sanitizar_texto(f"Fecha: {prop['fecha']} "), 'T R', 1, 'R', True)
        pdf.set_font('Helvetica', '', 10)
        pdf.set_text_color(*C_DARK)
        pdf.multi_cell(0, 6, sanitizar_texto(f"Propuesta:\n{prop['propuesta']}\n"), 'L R B', 'J', False)
        pdf.ln(3)

    shutil.rmtree(temp_dir, ignore_errors=True)
    return pdf.output(dest='S').encode('latin-1', 'replace')

def generar_pdf_rondines(df_resumen, escaneos_totales, alertas_fraude, fecha_str):
    temp_dir = os.path.join(os.path.dirname(__file__), f'temp_img_{uuid.uuid4().hex}')
    os.makedirs(temp_dir, exist_ok=True)
    
    plt.figure(figsize=(7, 4))
    plt.bar(df_resumen['Colaborador'], df_resumen['% Cumplimiento'], color=HEX_NAVY)
    plt.axhline(90, color='red', linestyle='--')
    plt.title('% Cumplimiento por Colaborador', fontsize=10, fontweight='bold')
    plt.tight_layout()
    p_bar = os.path.join(temp_dir, 'bar_ron.png')
    plt.savefig(p_bar)
    plt.close()

    pdf = SunhavenPDF()
    pdf.titulo_header = "REPORTE DE AUDITORÍA - RONDINES NOCTURNOS"
    pdf.cover_page("SUPERVISIÓN Y CUIDADO CONTINUO", "Auditoría Operativa", fecha_str)
    
    pdf.add_page()
    pdf.set_font('Helvetica', 'B', 14)
    pdf.set_text_color(*C_NAVY)
    pdf.cell(0, 8, sanitizar_texto("1. CUMPLIMIENTO OPERATIVO MENSUAL"), 0, 1, 'L')
    pdf.set_font('Helvetica', '', 10)
    pdf.set_text_color(*C_DARK)
    pdf.multi_cell(0, 6, sanitizar_texto("Esquema: 3 rondines obligatorios por noche. La siguiente tabla evalúa el cumplimiento según el rol asignado (Los días domingo están excluidos del cálculo de desempeño):"))
    pdf.ln(5)
    
    datos_tabla = [[r['Colaborador'], r['Turnos Meta'], r['Meta Rondas'], r['Rondas Real.'], f"{r['% Cumplimiento']:.1f}%"] for _, r in df_resumen.iterrows()]
    tabla_centrada(pdf, ["Colaborador", "Turnos Meta", "Meta Rondas", "Rondas Real.", "% Cumplimiento"], datos_tabla, [60, 30, 30, 30, 40])
    pdf.ln(10)
    
    pdf.set_font('Helvetica', 'B', 10)
    pdf.cell(0, 6, sanitizar_texto("Resumen de Escaneos y Auditoría Antifraude:"), 0, 1)
    pdf.set_font('Helvetica', '', 10)
    pdf.cell(0, 6, sanitizar_texto(f"Total de escaneos QR procesados en el periodo: {escaneos_totales}"), 0, 1)
    pdf.cell(0, 6, sanitizar_texto(f"Alertas de velocidad detectadas (posible fraude < 60s): {alertas_fraude}"), 0, 1)
    pdf.ln(5)
    
    pdf.image(p_bar, x=30, w=150)
    shutil.rmtree(temp_dir, ignore_errors=True)
    return pdf.output(dest='S').encode('latin-1', 'replace')

def generar_pdf_legal_bytes(categorias_dict, checks_dict, porcentaje):
    pdf = SunhavenPDF()
    pdf.titulo_header = "REPORTE EJECUTIVO - BLINDAJE NORMATIVO"
    pdf.cover_page("CUMPLIMIENTO LEGAL Y NORMATIVO", "Auditoría de Prevención Gubernamental", datetime.now().strftime("%d/%m/%Y"))
    pdf.add_page()
    pdf.set_font('Helvetica', 'B', 14)
    pdf.set_text_color(*C_NAVY)
    pdf.cell(0, 8, sanitizar_texto("ESTATUS GLOBAL DE BLINDAJE INSTITUCIONAL"), 0, 1, 'L')
    pdf.set_font('Helvetica', '', 11)
    pdf.set_text_color(*C_DARK)
    pdf.cell(0, 8, sanitizar_texto(f"Nivel de cumplimiento general de la institución: {porcentaje:.1f}%"), 0, 1)
    pdf.ln(5)
    for cat, items in categorias_dict.items():
        pdf.set_fill_color(220, 230, 240) 
        pdf.set_text_color(*C_NAVY)
        pdf.set_font('Helvetica', 'B', 8)
        pdf.cell(190, 7, sanitizar_texto(f" {cat.upper()}"), 1, 1, 'L', True)
        fill_row = False
        for req in items:
            estado = checks_dict.get(req, False)
            pdf.set_fill_color(*C_LIGHT) if fill_row else pdf.set_fill_color(255, 255, 255)
            pdf.set_text_color(*C_DARK)
            pdf.set_font('Helvetica', '', 7.5)
            req_texto = sanitizar_texto(f"   - {req}"[:115] + ("..." if len(req) > 112 else ""))
            pdf.cell(150, 7, req_texto, 1, 0, 'L', fill_row)
            pdf.set_text_color(39, 174, 96) if estado else pdf.set_text_color(231, 76, 60)
            pdf.set_font('Helvetica', 'B', 8)
            pdf.cell(40, 7, sanitizar_texto("CUMPLE" if estado else "PENDIENTE"), 1, 1, 'C', fill_row)
            fill_row = not fill_row
    return pdf.output(dest='S').encode('latin-1', 'replace')

# ==========================================
# 3. ETL Y CÁLCULOS
# ==========================================
@st.cache_data(ttl=600)
def cargar_datos_operaciones():
    try:
        gc = gspread.service_account_from_dict(json.loads(st.secrets["GOOGLE_JSON"])) if "GOOGLE_JSON" in st.secrets else gspread.service_account(filename=PATH_CREDS)
        dfs = {
            "n": pd.DataFrame(gc.open_by_url("https://docs.google.com/spreadsheets/d/10wWKmjsyj501OXaFWs7Rd_XF_2R-H0YzV66B2K6HvPE/edit").get_worksheet(0).get_all_records()),
            "v": pd.DataFrame(gc.open_by_url("https://docs.google.com/spreadsheets/d/1C1AVmNXG0ggRekB1HF4_IhX-NGiTkwzgvZn2Z31rCsc/edit").get_worksheet(0).get_all_records()),
            "s": pd.DataFrame(gc.open_by_url("https://docs.google.com/spreadsheets/d/1wdP3mbW_k4a90ubPG-ZQy8FvfAZiwKnhPjCj1BZHcBs/edit").get_worksheet(0).get_all_records())
        }
        for k in dfs: dfs[k].columns = dfs[k].columns.str.strip().str.replace('\n', ' ')
        return dfs
    except Exception as e: return None

@st.cache_data(ttl=600)
def fetch_kaizen_data():
    try:
        gc = gspread.service_account_from_dict(json.loads(st.secrets["GOOGLE_JSON"])) if "GOOGLE_JSON" in st.secrets else gspread.service_account(filename=PATH_CREDS)
        return pd.DataFrame(gc.open("SUNHAVEN_KAIZEN (Respuestas)").get_worksheet(0).get_all_records())
    except: return pd.DataFrame()

def cargar_bitacora():
    if os.path.exists(PATH_BITACORA): return pd.read_csv(PATH_BITACORA)
    df = pd.DataFrame(columns=["FECHA", "EMPLEADO", "INCIDENCIA", "OBSERVACION"])
    os.makedirs(os.path.dirname(PATH_BITACORA), exist_ok=True)
    df.to_csv(PATH_BITACORA, index=False)
    return df

def guardar_incidencia(fecha, empleado, incidencia, obs):
    df = cargar_bitacora()
    df = pd.concat([df, pd.DataFrame([{"FECHA": fecha, "EMPLEADO": empleado, "INCIDENCIA": incidencia, "OBSERVACION": obs}])], ignore_index=True)
    df.to_csv(PATH_BITACORA, index=False)

def borrar_incidencia(index):
    df = cargar_bitacora()
    df.drop(index).to_csv(PATH_BITACORA, index=False)

def limpiar_biometrico(file_bytes):
    ws = load_workbook(io.BytesIO(file_bytes), data_only=True).active
    datos, emp = [], None
    for row in ws.iter_rows(values_only=True):
        row_str = [str(cell) if cell is not None else "" for cell in row]
        if "ID:" in row_str[0]:
            try: emp = row_str[row_str.index("Nombre:") + 2].strip()
            except ValueError: pass
        elif emp and any(":" in str(cell) for cell in row_str):
            for dia, celda in enumerate(row_str, 1):
                celda = str(celda).strip()
                if len(celda) >= 5 and ":" in celda: datos.append({"Checador": emp, "Día": dia, "Entrada": celda[:5]})
            emp = None
    return pd.DataFrame(datos)

def procesar_super_nomina(df_bio, df_bitacora, df_kaizen, mes_num, anio_num):
    ret_list = []
    if not df_bio.empty:
        for _, row in df_bio.iterrows():
            ch = str(row['Checador']).upper()
            if ch in EMPLEADOS_DB:
                nm = EMPLEADOS_DB[ch]
                ent = row['Entrada']
                try:
                    he = datetime.strptime(ent, "%H:%M").time()
                    lim = HORA_ENTRADA_NOCHE if nm in ENFERMERAS_NOCHE else HORA_ENTRADA_DIA
                    if he > lim: ret_list.append({"FECHA": f"{anio_num}-{mes_num:02d}-{row['Día']:02d}", "EMPLEADO": nm, "INCIDENCIA": "Retardo Biométrico", "OBSERVACION": f"Entró a las {ent}"})
                except: pass
    df_ret = pd.DataFrame(ret_list)
    
    part, nopart, props = [], [], []
    stats_k = {'curr_si':0, 'curr_no':0, 'prev_si':0, 'prev_no':0, 'lista_no':[]}
    if not df_kaizen.empty:
        df_kaizen['Marca temporal'] = pd.to_datetime(df_kaizen['Marca temporal'], dayfirst=True, errors='coerce')
        df_k_curr = df_kaizen[(df_kaizen['Marca temporal'].dt.month == mes_num) & (df_kaizen['Marca temporal'].dt.year == anio_num)]
        part = df_k_curr.iloc[:, 1].str.strip().unique().tolist()
        nopart = [e for e in ENFERMERAS_LISTA if e not in part and e not in EXCEPCIONES_KAIZEN]
        
        m_prev = mes_num - 1 if mes_num > 1 else 12
        a_prev = anio_num if mes_num > 1 else anio_num - 1
        part_p = df_kaizen[(df_kaizen['Marca temporal'].dt.month == m_prev) & (df_kaizen['Marca temporal'].dt.year == a_prev)].iloc[:, 1].str.strip().unique().tolist()
        stats_k = {'curr_si': len([e for e in ENFERMERAS_LISTA if e in part]), 'curr_no': len(nopart), 'prev_si': len([e for e in ENFERMERAS_LISTA if e in part_p]), 'prev_no': len([e for e in ENFERMERAS_LISTA if e not in part_p and e not in EXCEPCIONES_KAIZEN]), 'lista_no': nopart}
        props = [{'nombre': r.iloc[1], 'fecha': r['Marca temporal'].strftime("%d/%m/%Y"), 'propuesta': str(r.iloc[2])} for _, r in df_k_curr.iterrows()]
            
    df_admin = pd.DataFrame([{"FECHA": f"{anio_num}-{mes_num:02d}-28", "EMPLEADO": e, "INCIDENCIA": "Falla Admin/Kaizen", "OBSERVACION": "No presentó propuesta"} for e in nopart])
    df_bitacora['FECHA'] = pd.to_datetime(df_bitacora['FECHA'], errors='coerce')
    df_bit = df_bitacora[(df_bitacora['FECHA'].dt.month == mes_num) & (df_bitacora['FECHA'].dt.year == anio_num)].copy()
    if not df_bit.empty: df_bit['FECHA'] = df_bit['FECHA'].dt.strftime('%Y-%m-%d')
    
    df_todas = pd.concat([df_ret, df_bit, df_admin], ignore_index=True)
    
    nomina = []
    for emp in sorted(list(set(EMPLEADOS_DB.values()))):
        df_e = df_todas[df_todas['EMPLEADO'] == emp]
        c_ret = len(df_e[df_e['INCIDENCIA'] == 'Retardo Biométrico'])
        f_kz = not df_e[df_e['INCIDENCIA'] == 'Falla Admin/Kaizen'].empty
        f_gr = len(df_e[df_e['INCIDENCIA'].str.contains('Grave', case=False, na=False)])
        
        bp = 500 if c_ret <= 3 else 0
        bu = 500
        ba = 0 if f_kz else 500
        if f_gr > 0: bp, bu, ba = 0, 0, 0
            
        nomina.append({"COLABORADOR": emp, "RETARDOS": c_ret, "INCIDENCIAS LEVES/GRAVES": len(df_e) - c_ret - (1 if f_kz else 0), "$ PUNTUAL": bp, "$ UNIFORM": bu, "$ ADMIN": ba, "TOTAL A PAGAR": bp + bu + ba})
        
    return pd.DataFrame(nomina), df_todas, df_ret, stats_k, props

# ==========================================
# 4. APLICACIÓN PRINCIPAL (ENRUTADOR)
# ==========================================
def main():
    with st.sidebar:
        st.markdown("### NAVEGADOR EMPRESARIAL")
        modulo_activo = st.radio("Seleccione el Módulo:", ["Dashboard de Operaciones", "Gestión de Nómina", "Turno Nocturno"], label_visibility="collapsed")
        st.divider()
        st.markdown("### FILTROS")
        
        fecha_inicio, fecha_fin, mes_eval, anio_eval, file_asis = None, None, None, None, None
        
        if modulo_activo in ["Dashboard de Operaciones", "Turno Nocturno"]:
            hoy = datetime.now()
            fechas = st.date_input("Rango de Análisis", [hoy - timedelta(days=30), hoy])
            if len(fechas) == 2: fecha_inicio, fecha_fin = fechas
            else: st.stop()
        else:
            mes_eval = st.selectbox("Mes", range(1, 13), index=datetime.now().month-1)
            anio_eval = st.number_input("Año", min_value=2020, max_value=2050, value=datetime.now().year)
            file_asis = st.file_uploader("Archivo Biométrico (.xlsx)", type=['xlsx', 'csv'])
            
        st.divider()
        if st.button("Sincronizar Datos DB", use_container_width=True):
            st.cache_data.clear()
            st.rerun()

    # ---------------------------------------------------------
    # MÓDULO 1: DASHBOARD OPERATIVO
    # ---------------------------------------------------------
    if modulo_activo == "Dashboard de Operaciones":
        st.title("Dashboard de Operaciones")
        data = cargar_datos_operaciones()
        if data is None: st.stop()
        df_ron, df_rop, df_serv = data["n"].copy(), data["v"].copy(), data["s"].copy()

        try:
            def traducir(s): return pd.to_numeric(s.astype(str).str.strip().str.lower().replace({'sí': 10, 'si': 10, 'no': 0, 'cumple': 10, 'no cumple': 0, 'separada': 10, 'bien': 10, 'ok': 10, 'limpio': 10, 'sucio': 0}), errors='coerce') * 10
            c_5s, c_cam = [c for c in df_rop.columns if "Orden" in c][0], [c for c in df_rop.columns if "Tendido" in c][0]
            c_enf = [c for c in df_rop.columns if "asignado" in c.lower() or "evaluad" in c.lower()]
            c_enf = c_enf[0] if c_enf else [c for c in df_rop.columns if "enfermera" in c.lower() or "nombre" in c.lower()][0]
            c_uni, c_bas, c_lav_r, c_lav_j, c_lim = [c for c in df_serv.columns if "uniforme" in c.lower()][0], [c for c in df_serv.columns if "basura" in c.lower()][0], [c for c in df_serv.columns if "ropa" in c.lower() or "separad" in c.lower()][0], [c for c in df_serv.columns if "jabón" in c.lower() or "jabon" in c.lower()][0], [c for c in df_serv.columns if "zonas asignadas" in c.lower()][0]

            for df in [df_ron, df_rop, df_serv]:
                df['Marca temporal'] = pd.to_datetime(df['Marca temporal'], dayfirst=True, errors='coerce')
                df['Fecha'] = df['Marca temporal'].dt.date
                
            df_ron = df_ron[(df_ron['Fecha'] >= fecha_inicio) & (df_ron['Fecha'] <= fecha_fin)]
            df_rop = df_rop[(df_rop['Fecha'] >= fecha_inicio) & (df_rop['Fecha'] <= fecha_fin)]
            df_serv = df_serv[(df_serv['Fecha'] >= fecha_inicio) & (df_serv['Fecha'] <= fecha_fin)]

            for c in [c_5s, c_cam]: df_rop[c] = traducir(df_rop[c])
            df_rop['Promedio'] = (df_rop[c_5s] + df_rop[c_cam]) / 2
            for c in [c_uni, c_bas, c_lav_r, c_lav_j, c_lim]: df_serv[c] = traducir(df_serv[c])

            t_A = sum(1 for d in pd.date_range(fecha_inicio, fecha_fin) if d.weekday() in [0, 2, 4])
            t_B = sum(1 for d in pd.date_range(fecha_inicio, fecha_fin) if d.weekday() in [1, 3, 5])
            df_ron['Hora'] = df_ron['Marca temporal'].dt.hour
            df_ron['Bloque'] = df_ron['Hora'].apply(lambda h: "B1" if 21<=h<=23 else "B2" if 0<=h<=3 else "B3" if 4<=h<=6 else None)
            
            rango_fechas = pd.date_range(fecha_inicio, fecha_fin)
            df_ron_val = df_ron[df_ron['Bloque'].notnull()].drop_duplicates(subset=['Fecha', 'Enfermera', 'Bloque'])
            conteo_noc = df_ron_val.groupby('Fecha').size()
            noc_diario = []
            for d in rango_fechas:
                exp = (len(ENFERMERAS_ROL_A)*3) if d.weekday() in [0,2,4] else (len(ENFERMERAS_ROL_B)*3 if d.weekday() in [1,3,5] else 0)
                noc_diario.append(min((conteo_noc.get(d.date(), 0)/exp)*100, 100) if exp>0 else None)
                
            df_evol_base = pd.DataFrame(index=rango_fechas.date)
            df_evol_base['Nocturno'] = noc_diario
            df_evol_base['Vespertino'] = df_evol_base.index.map(df_rop.groupby('Fecha')['Promedio'].mean())
            df_evol_base['Cocina'] = df_evol_base.index.map(df_serv.groupby('Fecha')[[c_uni, c_bas]].mean().mean(axis=1))
            df_evol_base['Lavandería'] = df_evol_base.index.map(df_serv.groupby('Fecha')[[c_lav_r, c_lav_j]].mean().mean(axis=1))
            df_evol_base['Limpieza'] = df_evol_base.index.map(df_serv.groupby('Fecha')[c_lim].mean())
            df_evol_base = df_evol_base.ffill().bfill().fillna(0)
            df_evol_base.index = pd.to_datetime(df_evol_base.index)

            p_noc = {enf: min((len(df_ron[(df_ron['Enfermera'] == enf) & (df_ron['Bloque'].notnull())][['Fecha', 'Bloque']].drop_duplicates()) / ((t_A if enf in ENFERMERAS_ROL_A else t_B)*3) * 100), 100) if ((t_A if enf in ENFERMERAS_ROL_A else t_B)*3) > 0 else 100 for enf in ENFERMERAS_NOCHE}
            v_noc = sum(p_noc.values()) / len(p_noc) if p_noc else 0

            kpi = {"Nocturno": v_noc, "Vespertino": df_rop['Promedio'].mean(), "Cocina": (df_serv[c_uni].mean() + df_serv[c_bas].mean()) / 2, "Lavandería": (df_serv[c_lav_r].mean() + df_serv[c_lav_j].mean()) / 2, "Limpieza": df_serv[c_lim].mean()}
            kpi = {k: (v if pd.notna(v) else 0) for k, v in kpi.items()}
            ico = sum(kpi.values()) / len(kpi)
            
            df_a = pd.DataFrame.from_dict(kpi, orient='index', columns=['V']).sort_values('V', ascending=False).reset_index()
            df_c = pd.DataFrame.from_dict({"Uniforme": df_serv[c_uni].mean(), "Basura": df_serv[c_bas].mean(), "Ropa": df_serv[c_lav_r].mean(), "Jabon": df_serv[c_lav_j].mean(), "Zonas": df_serv[c_lim].mean(), "5S Roperos": df_rop[c_5s].mean(), "Tendido Camas": df_rop[c_cam].mean(), "Rondines Noct": v_noc}, orient='index', columns=['V']).sort_values('V', ascending=True).reset_index()

            personal_diurno = set(df_rop[col_enf].dropna().unique()) - set(ENFERMERAS_NOCHE)
            ranking_data = [{"Colaborador": e, "Turno": "Nocturno", "Puntaje (%)": round(p_noc.get(e,0),1)} for e in ENFERMERAS_NOCHE if p_noc.get(e) is not None] + [{"Colaborador": e, "Turno": "Vespertino", "Puntaje (%)": round(df_rop[df_rop[col_enf]==e]['Promedio'].mean(),1)} for e in personal_diurno if pd.notna(df_rop[df_rop[col_enf]==e]['Promedio'].mean())]
            df_ranking = pd.DataFrame(ranking_data).sort_values("Puntaje (%)", ascending=False).reset_index(drop=True)

        except Exception as e:
            st.error(f"Error procesando operaciones: {e}")
            ico = 0

        st.markdown(f"""<div class='exec-header'>
            <div><p style='margin:0; font-weight:700; color:#64748b; font-size:12px; text-transform:uppercase;'>Estatus Institucional</p>
            <h2 style='margin:0; color:{HEX_GREEN if ico >= 90 else HEX_RED};'>{"ESTABLE" if ico >= 90 else "ATENCIÓN REQUERIDA"}</h2></div>
            <div style='text-align:right;'><p style='margin:0; font-weight:700; color:#64748b; font-size:12px; text-transform:uppercase;'>ICO Maestro</p>
            <h1 style='margin:0; font-size:48px;'>{ico:.1f}%</h1></div></div>""", unsafe_allow_html=True)

        tabs_op = st.tabs(["Tablero de Control", "Evolución Temporal", "Rendimiento Individual", "Blindaje Legal", "Data Cruda"])
        
        with tabs_op[0]:
            if st.button("Generar Reporte de Operaciones (PDF)", type="primary"):
                pdf_b = generar_pdf_dashboard_op(ico, "ESTABLE" if ico >= 90 else "ATENCIÓN REQUERIDA", df_a, df_c, df_evol_base, f"{fecha_inicio.strftime('%d/%m/%Y')} al {fecha_fin.strftime('%d/%m/%Y')}")
                st.download_button("Descargar Archivo", data=pdf_b, file_name="Reporte_Operaciones.pdf", mime="application/pdf")
            
            cols = st.columns(5)
            for i, (a, v) in enumerate(kpi.items()):
                color, flecha = (HEX_GREEN, '↑') if v >= 90 else (HEX_RED, '↓')
                with cols[i]: st.markdown(f"<div class='kpi-card'><p class='metric-label'>{a}</p><p class='metric-value' style='color:{color};'>{v:.1f}% {flecha}</p></div>", unsafe_allow_html=True)
            st.divider()
            c1, c2 = st.columns(2)
            with c1:
                st.write("### Pareto por Área")
                fig = px.bar(df_a, x='index', y='V', text_auto='.1f', color_discrete_sequence=[HEX_NAVY])
                fig.add_hline(y=90, line_dash="dash", line_color="red")
                st.plotly_chart(fig, use_container_width=True)
            with c2:
                st.write("### Análisis de Causa Raíz")
                fig_c = px.bar(df_c, x='V', y='index', orientation='h', text_auto='.1f', color_discrete_sequence=[HEX_RED])
                fig_c.add_vline(x=90, line_dash="dash", line_color="black")
                st.plotly_chart(fig_c, use_container_width=True)

        with tabs_op[1]:
            st.write("### Evolución Histórica de Todos los Departamentos")
            agrupacion = st.radio("Seleccionar Agrupación Temporal:", ["Día", "Semana", "Mes", "Año"], horizontal=True)
            
            if agrupacion == "Semana":
                df_plot = df_evol_base.resample('W-MON').mean()
                df_plot.index = df_plot.index.strftime('Semana %W - %Y')
            elif agrupacion == "Mes":
                df_plot = df_evol_base.resample('ME').mean()
                df_plot.index = df_plot.index.strftime('%Y-%m')
            elif agrupacion == "Año":
                df_plot = df_evol_base.resample('YE').mean()
                df_plot.index = df_plot.index.strftime('%Y')
            else:
                df_plot = df_evol_base.copy()
                df_plot.index = df_plot.index.strftime('%Y-%m-%d')

            fig_evol = px.line(df_plot, labels={"value": "Cumplimiento (%)", "index": "Periodo", "variable": "Área Operativa"}, markers=True)
            fig_evol.add_hline(y=90, line_dash="dot", line_color="red", annotation_text="Línea Base 90%")
            st.plotly_chart(fig_evol, use_container_width=True)

        with tabs_op[2]:
            st.write("### Rendimiento Detallado por Personal")
            st.dataframe(df_ranking, use_container_width=True)

        with tabs_op[3]:
            st.write("### Auditoría y Blindaje Institucional")
            cats = {
                "REGULACIÓN SANITARIA Y OPERATIVA - COPRISJAL/SSA": [
                    "[CRÍTICO - Anual] Aviso de Funcionamiento: Verificar documento vigente y exhibido.",
                    "[CRÍTICO - Anual] Responsable Sanitario: Validar aviso y nombramiento registrado.",
                    "[CRÍTICO - Diario] Medicamentos controlados: Verificar resguardo y control foliado.",
                    "[ALTO - Mensual] Accesibilidad NOM-031: Verificar rampas y barandales.",
                    "[CRÍTICO - Mensual] Expedientes clínicos: Validar integración y resguardo.",
                    "[CRÍTICO - Semanal] Manejo RPBI: Verificar contenedores y bolsas.",
                    "[ALTO - Diario] Higiene alimentaria: Control de temperaturas y limpieza en cocina."
                ],
                "SEGURIDAD Y SALUD LABORAL - STPS": [
                    "[ALTO - Anual] RIT: Verificar Reglamento Interior de Trabajo firmado.",
                    "[ALTO - Anual] NOM-035: Aplicar guía preventiva de riesgos psicosociales.",
                    "[ALTO - Trimestral] Comisión Mixta: Validar actas y recorridos de seguridad.",
                    "[ALTO - Semestral] Ergonomía (Movilización pacientes): Verificar capacitación y DC-3."
                ],
                "PROTECCIÓN CIVIL Y ECOLOGÍA MUNICIPAL": [
                    "[CRÍTICO - Anual] Programa Interno PIPC: Validar autorización vigente.",
                    "[CRÍTICO - Anual] Responsabilidad Civil: Verificar póliza de seguro vigente.",
                    "[CRÍTICO - Anual] Dictamen estructural: Validar dictamen DRO.",
                    "[CRÍTICO - Mensual] Extintores: Revisar vigencia, señalización y recarga.",
                    "[CRÍTICO - Mensual] Detectores de humo: Validar funcionamiento.",
                    "[ALTO - Mensual] Evacuación: Verificar rutas, luces y señalética.",
                    "[ALTO - Anual] Brigadas: Validar constancias de capacitación (DC-3).",
                    "[CRÍTICO - Semestral] Simulacros: Revisar bitácoras y formatos de evacuación."
                ],
                "CUMPLIMIENTO LEGAL Y PRIVACIDAD": [
                    "[ALTO - Permanente] Aviso de privacidad INAI: Verificar exhibición y anexos.",
                    "[CRÍTICO - Permanente] Datos sensibles: Confirmar consentimientos en expedientes.",
                    "[ALTO - Permanente] Contratos de servicios: Verificar contratos firmados y vigentes.",
                    "[ALTO - Anual] Contrato adhesión PROFECO: Confirmar registro vigente."
                ]
            }
            if 'checks' not in st.session_state: st.session_state.checks = {i: False for sub in cats.values() for i in sub}
            c_l1, c_l2 = st.columns([0.7, 0.3])
            with c_l1:
                for cat, items in cats.items():
                    with st.expander(cat, expanded=True):
                        for item in items: st.session_state.checks[item] = st.checkbox(item, value=st.session_state.checks[item], key=item)
            pct = (sum(st.session_state.checks.values()) / len(st.session_state.checks)) * 100
            with c_l2:
                st.write(f"#### Índice: {int(pct)}%")
                st.progress(pct / 100)
                if st.button("Generar Reporte Legal"): st.session_state['pl'] = generar_pdf_legal_bytes(cats, st.session_state.checks, pct)
                if 'pl' in st.session_state: st.download_button("Descargar Auditoría PDF", data=st.session_state['pl'], file_name=f"Legal_{datetime.now().strftime('%d%m%Y')}.pdf", mime="application/pdf")

        with tabs_op[4]:
            st.write("### Tuberías de Datos Operativas (Data Cruda)")
            sub1, sub2, sub3 = st.tabs(["Servicios Generales", "Enfermería Vespertina", "Rondines Nocturnos"])
            with sub1: st.dataframe(df_serv, use_container_width=True)
            with sub2: st.dataframe(df_rop, use_container_width=True)
            with sub3: st.dataframe(df_ron, use_container_width=True)

    # ---------------------------------------------------------
    # MÓDULO 2: GESTIÓN DE NÓMINA
    # ---------------------------------------------------------
    elif modulo_activo == "Gestión de Nómina":
        st.title("Gestión de Nómina y Mejora Continua")
        df_bitacora = cargar_bitacora()
        tabs_nom = st.tabs(["Dictamen de Nómina", "Bitácora Digital", "Data Cruda"])

        with tabs_nom[1]:
            st.write("### Registro Manual de Supervisión")
            with st.form("form_incidencia"):
                c_f1, c_f2, c_f3 = st.columns(3)
                with c_f1: f_fecha = st.date_input("Fecha")
                with c_f2: f_emp = st.selectbox("Colaborador", sorted(list(set(EMPLEADOS_DB.values()))))
                with c_f3: f_inc = st.selectbox("Incidencia", TIPO_INCIDENCIAS)
                f_obs = st.text_input("Observaciones")
                if st.form_submit_button("Guardar en Bitácora"):
                    guardar_incidencia(f_fecha.strftime('%Y-%m-%d'), f_emp, f_inc, f_obs)
                    st.rerun()
            st.dataframe(df_bitacora, use_container_width=True)
            if not df_bitacora.empty:
                b_idx = st.number_input("ID a borrar", 0, len(df_bitacora)-1, 0)
                if st.button("Eliminar"): borrar_incidencia(b_idx); st.rerun()

        with tabs_nom[0]:
            if file_asis:
                if st.button("Procesar Nómina vs Kaizen", type="primary"):
                    with st.spinner("Procesando..."):
                        st.session_state['nom'] = procesar_super_nomina(limpiar_biometrico(file_asis.read()), df_bitacora, fetch_kaizen_data(), mes_eval, anio_eval)
            
            if 'nom' in st.session_state:
                df_n, df_i, df_r, s_k, p_k = st.session_state['nom']
                m_str = f"{mes_eval:02d}/{anio_eval}"
                st.markdown(f"<h1 style='color:{HEX_GREEN};'>Total a Dispersar: ${df_n['TOTAL A PAGAR'].sum():,}</h1>", unsafe_allow_html=True)
                pdf_b = generar_pdf_nomina(df_n, df_i, df_r, s_k, p_k, m_str)
                st.download_button("Descargar Reporte de Nómina (PDF)", data=pdf_b, file_name=f"Nomina_{m_str.replace('/','_')}.pdf", mime="application/pdf")
                st.dataframe(df_n, use_container_width=True, hide_index=True)

        with tabs_nom[2]:
            if 'nom' in st.session_state:
                sub1, sub2 = st.tabs(["Retardos Biométricos", "Propuestas Kaizen"])
                with sub1: st.dataframe(st.session_state['nom'][2], use_container_width=True)
                with sub2: st.dataframe(fetch_kaizen_data(), use_container_width=True)

    # ---------------------------------------------------------
    # MÓDULO 3: TURNO NOCTURNO
    # ---------------------------------------------------------
    elif modulo_activo == "Turno Nocturno":
        st.title("Evaluación de Turno Nocturno")
        data = cargar_datos_operaciones()
        if data is None: st.stop()
        
        df_ron = data["n"].copy()
        df_ron['Marca temporal'] = pd.to_datetime(df_ron['Marca temporal'], dayfirst=True, errors='coerce')
        df_ron['Fecha'] = df_ron['Marca temporal'].dt.date
        df_ron = df_ron[(df_ron['Fecha'] >= fecha_inicio) & (df_ron['Fecha'] <= fecha_fin)]

        t_A = sum(1 for d in pd.date_range(fecha_inicio, fecha_fin) if d.weekday() in [0, 2, 4] and d.weekday() != 6)
        t_B = sum(1 for d in pd.date_range(fecha_inicio, fecha_fin) if d.weekday() in [1, 3, 5] and d.weekday() != 6)
        
        df_ron['Hora'] = df_ron['Marca temporal'].dt.hour
        df_ron['Bloque'] = df_ron['Hora'].apply(lambda h: "B1" if 21<=h<=23 else "B2" if 0<=h<=3 else "B3" if 4<=h<=6 else None)
        df_valido = df_ron[df_ron['Bloque'].notnull()].drop_duplicates(subset=['Fecha', 'Enfermera', 'Bloque'])

        df_ron_sort = df_ron.sort_values(by=['Enfermera', 'Marca temporal'])
        df_ron_sort['Diff'] = df_ron_sort.groupby('Enfermera')['Marca temporal'].diff().dt.total_seconds()
        alertas = len(df_ron_sort[df_ron_sort['Diff'] < 60])
        escaneos = len(df_ron)

        datos_noc = []
        for enf in ENFERMERAS_NOCHE:
            turnos_m = t_A if enf in ENFERMERAS_ROL_A else t_B
            rondas_m = turnos_m * 3
            rondas_r = len(df_valido[df_valido['Enfermera'] == enf])
            pct = min((rondas_r / rondas_m * 100), 100) if rondas_m > 0 else 100
            datos_noc.append({"Colaborador": enf, "Turnos Meta": turnos_m, "Meta Rondas": rondas_m, "Rondas Real.": rondas_r, "% Cumplimiento": pct})
        
        df_resumen = pd.DataFrame(datos_noc)
        v_noc = df_resumen['% Cumplimiento'].mean()

        st.markdown(f"""<div class='exec-header'>
            <div><p style='margin:0; font-weight:700; color:#64748b; font-size:12px; text-transform:uppercase;'>Cumplimiento de Seguridad</p>
            <h2 style='margin:0; color:{HEX_GREEN if v_noc >= 90 else HEX_RED};'>{"ESTABLE" if v_noc >= 90 else "ALERTA"}</h2></div>
            <div style='text-align:right;'><p style='margin:0; font-weight:700; color:#64748b; font-size:12px; text-transform:uppercase;'>Promedio Global Nocturno</p>
            <h1 style='margin:0; font-size:48px;'>{v_noc:.1f}%</h1></div></div>""", unsafe_allow_html=True)

        tabs_noc = st.tabs(["Auditoría de Rondas", "Log Antifraude"])
        with tabs_noc[0]:
            if st.button("Generar Reporte Rondines (PDF)", type="primary"):
                pdf_b = generar_pdf_rondines(df_resumen, escaneos, alertas, f"{fecha_inicio.strftime('%d/%m/%Y')} al {fecha_fin.strftime('%d/%m/%Y')}")
                st.download_button("Descargar Archivo", data=pdf_b, file_name="Reporte_Rondines.pdf", mime="application/pdf")
            st.dataframe(df_resumen, use_container_width=True, hide_index=True)
        with tabs_noc[1]:
            st.write(f"**Escaneos Totales:** {escaneos} | **Alertas Fraude (<60s):** {alertas}")
            st.dataframe(df_ron_sort[['Marca temporal', 'Enfermera', 'Diff']].sort_values(by="Marca temporal", ascending=False), use_container_width=True)

if __name__ == "__main__":
    main()